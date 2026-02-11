from datetime import date
from io import BytesIO
from unittest.mock import patch
import uuid

from django.test import TestCase

from infraohjelmointi_api.models import (
    Project,
    TalpaProjectOpening,
    TalpaProjectType,
    TalpaServiceClass,
    TalpaAssetClass,
    TalpaProjectNumberRange,
    Person,
    ProjectType,
    ProjectPhase,
    ProjectCategory,
    ProjectClass,
)
from infraohjelmointi_api.services import TalpaExcelService
from openpyxl import load_workbook


class TalpaExcelServiceTestCase(TestCase):
    def setUp(self):
        self.service = TalpaExcelService()
        
        self.project_type = ProjectType.objects.create(
            id=uuid.UUID("844e3102-7fb0-453b-ad7b-cf69b1644166"),
            value="Test Project Type"
        )
        
        self.project_phase = ProjectPhase.objects.create(
            id=uuid.UUID("081ff330-5b0a-4ddc-b39b-cd9e53070256"),
            value="Test Phase"
        )
        
        self.project_category = ProjectCategory.objects.create(
            id=uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef1234567890"),
            value="Test"
        )
        
        self.project_class = ProjectClass.objects.create(
            id=uuid.UUID("b2c3d4e5-f6a7-8901-bcde-f12345678901"),
            name="Test Class",
            path="8/8 01",
        )
        
        self.project = Project.objects.create(
            id=uuid.UUID("ffef67fa-a790-4e22-b979-383e7da62374"),
            name="Test Project",
            description="Test description",
            type=self.project_type,
            phase=self.project_phase,
            category=self.project_category,
            projectClass=self.project_class,
        )
        
        self.talpa_project_type, _ = TalpaProjectType.objects.get_or_create(
            code="INV",
            priority="03",
            defaults={
                "name": "Infrastructure Investment",
                "category": "INVESTMENT",
            }
        )
        
        self.talpa_service_class, _ = TalpaServiceClass.objects.get_or_create(
            code="4601",
            defaults={
                "name": "Test Service",
                "projectTypePrefix": "2814I",
            }
        )
        
        self.talpa_asset_class, _ = TalpaAssetClass.objects.get_or_create(
            componentClass="K1230000",
            account="1230000",
            defaults={
                "name": "Test Asset",
                "holdingPeriodYears": 10,
            }
        )
        
        self.talpa_range, _ = TalpaProjectNumberRange.objects.get_or_create(
            projectTypePrefix="2814I",
            rangeStart="2814I01000",
            rangeEnd="2814I01599",
            defaults={
                "budgetAccount": "8 03 01 01",
                "budgetAccountNumber": "8030101",
            }
        )

    def test_generate_excel_creates_valid_file(self):
        opening = TalpaProjectOpening.objects.create(
            project=self.project,
            subject="Uusi",
            projectName="Test Project Name",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_range,
            budgetAccount="8 03 01 01 Test Budget",
            templateProject="2814I00000",
            projectStartDate=date(2026, 1, 15),
            projectEndDate=date(2032, 12, 31),
            streetAddress="Test Street 1",
            postalCode="00100",
            responsiblePerson="Test Person",
            serviceClass=self.talpa_service_class,
            assetClass=self.talpa_asset_class,
            profileName="Test Profile",
            investmentProfile="Z12550",
            readiness="Kesken",
        )
        
        excel_file = self.service.generate_excel(opening)
        
        self.assertIsInstance(excel_file, BytesIO)
        self.assertGreater(len(excel_file.getvalue()), 0)
        
        excel_file.seek(0)
        wb = load_workbook(excel_file)
        ws = wb.active
        
        self.assertEqual(ws.title, "Projektin avauslomake")
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.max_column, 15)

    def test_generate_excel_has_all_columns(self):
        opening = TalpaProjectOpening.objects.create(
            project=self.project,
            subject="Uusi",
            projectName="Test",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_range,
        )
        
        excel_file = self.service.generate_excel(opening)
        excel_file.seek(0)
        wb = load_workbook(excel_file)
        ws = wb.active
        
        expected_headers = [
            "Talousarviokohdan numero",
            "Talousarviokohdan nimi",
            "Projektinumeroväli",
            "Malliprojekti",
            "Laji",
            "Prioriteetti",
            "SAP nimi",
            "Osoite+postinumero",
            "Projekti alkaa -päättyy, huom.takuuaika",
            "Vastuuhenkilö",
            "Palveluluokka",
            "Käyttöomaisuusluokat",
            "Invest. profiili",
            "Profiilin nimi",
            "Valmius",
        ]
        
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, expected_headers)

    def test_generate_excel_date_formatting(self):
        opening = TalpaProjectOpening.objects.create(
            project=self.project,
            subject="Uusi",
            projectName="Test",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_range,
            projectStartDate=date(2026, 3, 15),
            projectEndDate=date(2032, 11, 30),
        )
        
        excel_file = self.service.generate_excel(opening)
        excel_file.seek(0)
        wb = load_workbook(excel_file)
        ws = wb.active
        
        schedule = ws["I2"].value
        
        self.assertEqual(schedule, "15.03.2026 - 30.11.2032")

    def test_generate_excel_computed_fields(self):
        opening = TalpaProjectOpening.objects.create(
            project=self.project,
            subject="Uusi",
            projectName="Test",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_range,
            budgetAccount="8 03 01 02 Perusparantaminen",
            serviceClass=self.talpa_service_class,
            assetClass=self.talpa_asset_class,
            streetAddress="Main Street 5",
            postalCode="00100",
        )
        
        excel_file = self.service.generate_excel(opening)
        excel_file.seek(0)
        wb = load_workbook(excel_file)
        ws = wb.active
        
        budget_num = ws["A2"].value
        budget_name = ws["B2"].value
        number_range = ws["C2"].value
        type_code = ws["E2"].value
        type_priority = ws["F2"].value
        address_full = ws["H2"].value
        service_code = ws["K2"].value
        asset_component = ws["L2"].value
        
        self.assertEqual(budget_num, "8 03 01 02")
        self.assertEqual(budget_name, "Perusparantaminen")
        self.assertEqual(number_range, "2814I01000 - 2814I01599")
        self.assertEqual(type_code, "INV")
        self.assertEqual(type_priority, "03")
        self.assertEqual(address_full, "Main Street 5 00100")
        self.assertEqual(service_code, "4601")
        self.assertEqual(asset_component, "K1230000")

    def test_generate_excel_empty_fields(self):
        opening = TalpaProjectOpening.objects.create(
            project=self.project,
            subject="Uusi",
            projectName="Test",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_range,
        )
        
        excel_file = self.service.generate_excel(opening)
        excel_file.seek(0)
        wb = load_workbook(excel_file)
        ws = wb.active
        
        self.assertIn(ws["H2"].value, [None, ""])
        self.assertIn(ws["J2"].value, [None, ""])

    def test_generate_batch_excel(self):
        project2 = Project.objects.create(
            name="Test Project 2",
            description="Test description 2",
            type=self.project_type,
            phase=self.project_phase,
            category=self.project_category,
            projectClass=self.project_class,
        )
        
        opening1 = TalpaProjectOpening.objects.create(
            project=self.project,
            subject="Uusi",
            projectName="Test 1",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_range,
        )
        
        opening2 = TalpaProjectOpening.objects.create(
            project=project2,
            subject="Muutos",
            projectName="Test 2",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_range,
        )
        
        excel_file = self.service.generate_batch_excel([opening1, opening2])
        excel_file.seek(0)
        wb = load_workbook(excel_file)
        ws = wb.active
        
        self.assertEqual(ws.max_row, 3)
        self.assertEqual(ws["G2"].value, "Test 1")
        self.assertEqual(ws["G3"].value, "Test 2")

    def test_get_filename_with_project_name(self):
        opening = TalpaProjectOpening.objects.create(
            project=self.project,
            subject="Uusi",
            projectName="Test",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_range,
        )
        
        filename = self.service.get_filename(opening)
        
        self.assertEqual(filename, "Talpa_avauslomake_Test Project.xlsx")

    def test_get_filename_format(self):
        opening = TalpaProjectOpening.objects.create(
            project=self.project,
            subject="Uusi",
            projectName="Test",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_range,
        )
        
        filename = self.service.get_filename(opening)
        
        self.assertTrue(filename.startswith("Talpa_avauslomake_"))
        self.assertTrue(filename.endswith(".xlsx"))
        self.assertIn("Test Project", filename)

