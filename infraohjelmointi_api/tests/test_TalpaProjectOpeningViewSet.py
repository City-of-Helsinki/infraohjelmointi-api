from datetime import date
from io import BytesIO
from unittest.mock import patch
import uuid

from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework import status
from openpyxl import load_workbook

from infraohjelmointi_api.models import (
    Project,
    TalpaProjectOpening,
    TalpaProjectType,
    TalpaServiceClass,
    TalpaAssetClass,
    TalpaProjectNumberRange,
    Person,
    ProjectType,
    ProjectPhase,
    ProjectCategory,
    ProjectClass,
)
from infraohjelmointi_api.views.BaseViewSet import BaseViewSet


@patch.object(BaseViewSet, "authentication_classes", new=[])
@patch.object(BaseViewSet, "permission_classes", new=[])
class TalpaProjectOpeningViewSetTestCase(TestCase):
    """Essential unit tests for TalpaProjectOpeningViewSet"""

    def setUp(self):
        """Set up test data"""
        self.client = APIClient()

        # Create test Person
        self.person = Person.objects.create(
            id=uuid.UUID("2c6dece3-cf93-45ba-867d-8f1dd14923fc"),
            firstName="Test",
            lastName="User",
            email="test@example.com"
        )

        # Create test ProjectType
        self.project_type = ProjectType.objects.create(
            id=uuid.UUID("844e3102-7fb0-453b-ad7b-cf69b1644166"),
            value="Test Project Type"
        )

        # Create test ProjectPhase
        self.project_phase = ProjectPhase.objects.create(
            id=uuid.UUID("081ff330-5b0a-4ddc-b39b-cd9e53070256"),
            value="Test Phase"
        )

        # Create test ProjectCategory
        self.project_category = ProjectCategory.objects.create(
            id=uuid.UUID("dbc92a70-8a8a-4a25-8014-14c7d16eb86c"),
            value="K5.1"
        )

        # Create test ProjectClass
        self.project_class = ProjectClass.objects.create(
            id=uuid.UUID("5f65a339-b3c9-48ee-a9b9-cb177546c241"),
            name="Test Class"
        )

        # Create test Project
        self.project = Project.objects.create(
            id=uuid.UUID("33814e76-7bdc-47c2-bf08-7ed43a96e042"),
            name="Test Project",
            description="Test description",
            type=self.project_type,
            phase=self.project_phase,
            category=self.project_category,
            projectClass=self.project_class,
        )

        # Create test TalpaProjectType
        self.talpa_project_type = TalpaProjectType.objects.create(
            code="8 03 01 01",
            name="Katujen uudisrakentaminen",
            category="KADUT, LIIKENNEVÄYLÄT JA RADAT",
            priority="Normaali",
            isActive=True
        )

        # Create test TalpaServiceClass
        self.talpa_service_class = TalpaServiceClass.objects.create(
            code="4601",
            name="Kadut ja yleiset alueet",
            projectTypePrefix="2814I",
            isActive=True
        )

        # Create test TalpaAssetClass
        self.talpa_asset_class = TalpaAssetClass.objects.create(
            componentClass="8103000",
            account="103000",
            name="Maa- ja vesialueet",
            category="Kiinteät rakenteet ja laitteet",
            isActive=True
        )

        # Create test TalpaProjectNumberRange
        self.talpa_project_number_range = TalpaProjectNumberRange.objects.create(
            projectTypePrefix="2814I",
            budgetAccount="8 03 01 01",
            budgetAccountNumber="2814100000",
            rangeStart="2814100003",
            rangeEnd="2814100300",
            majorDistrict="01",
            majorDistrictName="Eteläinen",
            area="011 Keskusta",
            isActive=True
        )

    def test_create_talpa_opening(self):
        """Test creating a new TalpaProjectOpening"""
        url = "/talpa-project-opening/"
        data = {
            "project": str(self.project.id),
            "priority": "Normaali",
            "subject": "Uusi",
            "projectName": "Test Talpa Project",
            "projectTypeId": str(self.talpa_project_type.id),
            "budgetAccount": "2814100000",
            "majorDistrict": "01 Eteläinen",
            "area": "011 Keskusta",
            "projectDescription": "Test description",
            "responsiblePerson": "John Doe",
            "responsiblePersonEmail": "john@example.com",
            "responsiblePersonPhone": "+358501234567",
            "serviceClassId": str(self.talpa_service_class.id),
            "assetClassId": str(self.talpa_asset_class.id),
        }

        response = self.client.post(url, data, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(TalpaProjectOpening.objects.count(), 1)

        talpa_opening = TalpaProjectOpening.objects.get(project=self.project)
        self.assertEqual(talpa_opening.projectName, "Test Talpa Project")
        self.assertEqual(talpa_opening.status, "excel_generated")
        self.assertEqual(talpa_opening.priority, "Normaali")
        self.assertEqual(talpa_opening.subject, "Uusi")
        self.assertFalse(talpa_opening.is_locked)

    def test_get_talpa_opening_by_id(self):
        """Test retrieving a TalpaProjectOpening by ID"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Test Project"
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["projectName"], "Test Project")
        self.assertEqual(response.data["status"], "excel_generated")
        self.assertFalse(response.data["isLocked"])

    def test_get_talpa_opening_by_project(self):
        """Test custom action to get TalpaProjectOpening by project ID"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Test Project"
        )

        url = f"/talpa-project-opening/by-project/{self.project.id}/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["projectName"], "Test Project")
        self.assertEqual(response.data["id"], str(talpa_opening.id))

    def test_get_talpa_opening_by_project_not_found(self):
        """Test getting TalpaProjectOpening by project ID when it doesn't exist"""
        url = f"/talpa-project-opening/by-project/{self.project.id}/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_update_talpa_opening(self):
        """Test updating a TalpaProjectOpening"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Original Name"
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/"
        data = {
            "projectName": "Updated Name",
            "projectDescription": "Updated description"
        }

        response = self.client.patch(url, data, format="json")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        talpa_opening.refresh_from_db()
        self.assertEqual(talpa_opening.projectName, "Updated Name")
        self.assertEqual(talpa_opening.projectDescription, "Updated description")

    def test_send_to_talpa_locks_form(self):
        """Test that sending to Talpa locks the form"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Test Project",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_project_number_range,
            status="excel_generated"
        )

        url = f"/api/talpa-project-opening/{talpa_opening.id}/send-to-talpa/"
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        talpa_opening.refresh_from_db()
        self.assertEqual(talpa_opening.status, "sent_to_talpa")
        self.assertTrue(talpa_opening.is_locked)
        self.assertTrue(response.data["isLocked"])

    def test_send_to_talpa_already_sent(self):
        """Test that sending to Talpa twice returns error"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Test Project",
            status="sent_to_talpa"
        )

        url = f"/api/talpa-project-opening/{talpa_opening.id}/send-to-talpa/"
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Already sent", response.data["detail"])

    def test_update_locked_form_returns_403(self):
        """Test that updating a locked form returns 403"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Test Project",
            status="sent_to_talpa"
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/"
        data = {"projectName": "Updated Name"}

        response = self.client.patch(url, data, format="json")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn("locked", response.data["detail"].lower())

    def test_get_priorities(self):
        """Test getting priority choices"""
        url = "/talpa-project-opening/priorities/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 2)
        priorities = [item["value"] for item in response.data]
        self.assertIn("Normaali", priorities)
        self.assertIn("Korkea", priorities)

    def test_get_subjects(self):
        """Test getting subject choices"""
        url = "/talpa-project-opening/subjects/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 3)
        subjects = [item["value"] for item in response.data]
        self.assertIn("Uusi", subjects)
        self.assertIn("Muutos", subjects)
        self.assertIn("Lukitus", subjects)

    def test_is_locked_property(self):
        """Test is_locked property on model"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            status="excel_generated"
        )
        self.assertFalse(talpa_opening.is_locked)

        talpa_opening.status = "sent_to_talpa"
        talpa_opening.save()
        self.assertTrue(talpa_opening.is_locked)

        # project_number_opened should also be locked
        talpa_opening.status = "project_number_opened"
        talpa_opening.save()
        self.assertTrue(talpa_opening.is_locked)

    def test_default_status_is_excel_generated(self):
        """Test that default status is excel_generated"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi"
        )

        self.assertEqual(talpa_opening.status, "excel_generated")

    def test_foreign_key_relationships(self):
        """Test that foreign key relationships work correctly"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectType=self.talpa_project_type,
            serviceClass=self.talpa_service_class,
            assetClass=self.talpa_asset_class
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsNotNone(response.data["projectType"])
        self.assertEqual(response.data["projectType"]["code"], "8 03 01 01")
        self.assertIsNotNone(response.data["serviceClass"])
        self.assertEqual(response.data["serviceClass"]["code"], "4601")
        self.assertIsNotNone(response.data["assetClass"])
        self.assertEqual(response.data["assetClass"]["componentClass"], "8103000")

    def test_create_with_new_address_fields(self):
        """Test creating TalpaProjectOpening with new address fields"""
        url = "/talpa-project-opening/"
        data = {
            "project": str(self.project.id),
            "priority": "Normaali",
            "subject": "Uusi",
            "streetAddress": "Testikatu 1",
            "postalCode": "00100",
        }

        response = self.client.post(url, data, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        talpa_opening = TalpaProjectOpening.objects.get(project=self.project)
        self.assertEqual(talpa_opening.streetAddress, "Testikatu 1")
        self.assertEqual(talpa_opening.postalCode, "00100")

    def test_create_with_schedule_fields(self):
        """Test creating with schedule date fields"""
        url = "/talpa-project-opening/"
        data = {
            "project": str(self.project.id),
            "priority": "Normaali",
            "subject": "Uusi",
            "projectStartDate": "2025-01-01",
            "projectEndDate": "2030-12-31",
        }
        
        response = self.client.post(url, data, format="json")
        
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        talpa_opening = TalpaProjectOpening.objects.get(project=self.project)
        self.assertEqual(str(talpa_opening.projectStartDate), "2025-01-01")
        self.assertEqual(str(talpa_opening.projectEndDate), "2030-12-31")

    def test_response_includes_schedule_fields(self):
        """Test that response includes schedule date fields"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectStartDate=date(2025, 1, 1),
            projectEndDate=date(2030, 12, 31),
        )
        
        url = f"/talpa-project-opening/{talpa_opening.id}/"
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("projectStartDate", response.data)
        self.assertIn("projectEndDate", response.data)
        self.assertEqual(response.data["projectStartDate"], "2025-01-01")
        self.assertEqual(response.data["projectEndDate"], "2030-12-31")

    def test_create_with_readiness_field(self):
        """Test creating with readiness field"""
        url = "/talpa-project-opening/"
        data = {
            "project": str(self.project.id),
            "priority": "Normaali",
            "subject": "Uusi",
            "readiness": "Kesken",
        }

        response = self.client.post(url, data, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        talpa_opening = TalpaProjectOpening.objects.get(project=self.project)
        self.assertEqual(talpa_opening.readiness, "Kesken")

    def test_holding_time_computed_field(self):
        """Test that holdingTime is computed from assetClass"""
        self.talpa_asset_class.holdingPeriodYears = 20
        self.talpa_asset_class.save()

        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            assetClass=self.talpa_asset_class
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["holdingTime"], 20)

    def test_create_with_asset_class_id(self):
        """Test that assetClassId works correctly"""
        url = "/talpa-project-opening/"
        data = {
            "project": str(self.project.id),
            "priority": "Normaali",
            "subject": "Uusi",
            "assetClassId": str(self.talpa_asset_class.id),
        }
        
        response = self.client.post(url, data, format="json")
        
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        talpa_opening = TalpaProjectOpening.objects.get(project=self.project)
        self.assertEqual(talpa_opening.assetClass, self.talpa_asset_class)

    # =========================================================================
    # Project Number Range Tests
    # =========================================================================

    def test_create_with_project_number_range_id(self):
        """Test creating TalpaProjectOpening with projectNumberRangeId"""
        url = "/talpa-project-opening/"
        data = {
            "project": str(self.project.id),
            "subject": "Uusi",
            "projectName": "Test Project",
            "projectNumberRangeId": str(self.talpa_project_number_range.id),
        }

        response = self.client.post(url, data, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        talpa_opening = TalpaProjectOpening.objects.get(project=self.project)
        self.assertEqual(talpa_opening.projectNumberRange, self.talpa_project_number_range)

    def test_response_includes_project_number_range_nested(self):
        """Test that GET response includes nested projectNumberRange data"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectNumberRange=self.talpa_project_number_range
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsNotNone(response.data["projectNumberRange"])
        self.assertEqual(response.data["projectNumberRange"]["rangeStart"], "2814100003")
        self.assertEqual(response.data["projectNumberRange"]["rangeEnd"], "2814100300")
        self.assertEqual(response.data["projectNumberRange"]["projectTypePrefix"], "2814I")

    # =========================================================================
    # Delete Protection Tests
    # =========================================================================

    def test_delete_unlocked_form_succeeds(self):
        """Test that deleting an unlocked form works"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            status="excel_generated"
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/"
        response = self.client.delete(url)

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertEqual(TalpaProjectOpening.objects.count(), 0)

    def test_delete_locked_form_returns_403(self):
        """Test that deleting a locked form returns 403"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            status="sent_to_talpa"
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/"
        response = self.client.delete(url)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn("locked", response.data["detail"].lower())
        # Verify it wasn't deleted
        self.assertEqual(TalpaProjectOpening.objects.count(), 1)

    # =========================================================================
    # Send to Talpa Validation Tests
    # =========================================================================

    def test_send_to_talpa_with_missing_fields_returns_400(self):
        """Test that sending incomplete form to Talpa returns 400 with missing fields"""
        # Create opening without required fields (projectName, projectType, projectNumberRange)
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            status="excel_generated"
            # Missing: projectName, projectType, projectNumberRange
        )

        url = f"/api/talpa-project-opening/{talpa_opening.id}/send-to-talpa/"
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("missing_fields", response.data)
        self.assertIn("SAP nimi", response.data["missing_fields"])
        self.assertIn("Laji", response.data["missing_fields"])
        self.assertIn("Projektinumeroväli", response.data["missing_fields"])

    def test_send_to_talpa_with_complete_form_succeeds(self):
        """Test that sending complete form to Talpa succeeds"""
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Complete Project",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_project_number_range,
            status="excel_generated"
        )

        url = f"/api/talpa-project-opening/{talpa_opening.id}/send-to-talpa/"
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        talpa_opening.refresh_from_db()
        self.assertEqual(talpa_opening.status, "sent_to_talpa")
        self.assertTrue(talpa_opening.is_locked)

    def test_unknown_field_ignored(self):
        """Test that unknown fields are ignored (e.g., old alias fields)"""
        url = "/talpa-project-opening/"
        data = {
            "project": str(self.project.id),
            "subject": "Uusi",
            "unknownField": "some value",  # Unknown field should be ignored
            "projectStartDate": "2025-01-01",
        }
        
        response = self.client.post(url, data, format="json")
        
        # Should succeed - unknown fields are ignored
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        talpa_opening = TalpaProjectOpening.objects.get(project=self.project)
        self.assertEqual(str(talpa_opening.projectStartDate), "2025-01-01")

    # =========================================================================
    # Default Priority Test
    # =========================================================================

    def test_default_priority_is_normaali(self):
        """Test that default priority is 'Normaali' when not provided"""
        url = "/talpa-project-opening/"
        data = {
            "project": str(self.project.id),
            "subject": "Uusi",
            # No priority provided
        }

        response = self.client.post(url, data, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        talpa_opening = TalpaProjectOpening.objects.get(project=self.project)
        self.assertEqual(talpa_opening.priority, "Normaali")

    # =========================================================================
    # Excel Download Tests
    # =========================================================================

    def test_download_excel_returns_file(self):
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Test Project",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_project_number_range,
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/download-excel/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])
        self.assertGreater(len(response.content), 0)

    def test_download_excel_updates_status(self):
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Test Project",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_project_number_range,
            status="excel_generated",
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/download-excel/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        talpa_opening.refresh_from_db()
        self.assertEqual(talpa_opening.status, "excel_generated")

    def test_download_excel_sets_status_to_excel_generated(self):
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Test Project",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_project_number_range,
            status="sent_to_talpa",
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/download-excel/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        talpa_opening.refresh_from_db()
        self.assertEqual(talpa_opening.status, "sent_to_talpa")

    def test_download_excel_valid_excel_content(self):
        talpa_opening = TalpaProjectOpening.objects.create(
            project=self.project,
            priority="Normaali",
            subject="Uusi",
            projectName="Test Project",
            projectType=self.talpa_project_type,
            projectNumberRange=self.talpa_project_number_range,
            projectStartDate=date(2026, 1, 15),
            projectEndDate=date(2032, 12, 31),
        )

        url = f"/talpa-project-opening/{talpa_opening.id}/download-excel/"
        response = self.client.get(url)

        excel_data = BytesIO(response.content)
        wb = load_workbook(excel_data)
        ws = wb.active

        self.assertEqual(ws.title, "Projektin avauslomake")
        self.assertEqual(ws["F2"].value, "Test Project")
        self.assertEqual(ws["G2"].value, "15.01.2026")
        self.assertEqual(ws["H2"].value, "31.12.2032")

