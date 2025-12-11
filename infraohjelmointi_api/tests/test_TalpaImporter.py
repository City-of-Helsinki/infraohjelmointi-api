"""
Unit tests for talpaimporter management command.

Tests the parsing functions for different Excel file formats.
"""
from django.test import TestCase
from django.core.management import call_command
from unittest.mock import MagicMock
import openpyxl
import tempfile
import os
import shutil

from infraohjelmointi_api.management.commands.talpaimporter import Command
from infraohjelmointi_api.models import TalpaProjectNumberRange, TalpaServiceClass, TalpaAssetClass


class TalpaImporterParsingTestCase(TestCase):
    """Test the parsing functions of talpaimporter"""

    def setUp(self):
        """Set up test fixtures"""
        self.command = Command()
        self.command.stdout = MagicMock()
        self.command.style = MagicMock()

    # =========================================================================
    # SAP Range Row Parsing Tests
    # =========================================================================

    def test_parse_sap_range_row_2814I_with_district(self):
        """Test parsing a 2814I range with district from SAP format"""
        # Mock row: [Cell(ETELÄINEN SUURPIIRI), Cell(2814I00003-2814I00300)]
        row = [
            MagicMock(value="ETELÄINEN SUURPIIRI"),
            MagicMock(value="2814I00003-2814I00300"),
        ]

        result = self.command._parse_sap_range_row(row)

        self.assertIsNotNone(result)
        self.assertEqual(result['projectTypePrefix'], '2814I')
        self.assertEqual(result['rangeStart'], '2814I00003')
        self.assertEqual(result['rangeEnd'], '2814I00300')
        self.assertEqual(result['majorDistrict'], '01')
        self.assertEqual(result['majorDistrictName'], 'Eteläinen suurpiiri')

    def test_parse_sap_range_row_2814I_koillinen(self):
        """Test parsing a 2814I range for Koillinen suurpiiri"""
        row = [
            MagicMock(value="KOILLINEN SUURPIIRI"),
            MagicMock(value="2814I01501-2814I02900"),
        ]

        result = self.command._parse_sap_range_row(row)

        self.assertIsNotNone(result)
        self.assertEqual(result['projectTypePrefix'], '2814I')
        self.assertEqual(result['rangeStart'], '2814I01501')
        self.assertEqual(result['rangeEnd'], '2814I02900')
        self.assertEqual(result['majorDistrict'], '05')
        self.assertEqual(result['majorDistrictName'], 'Koillinen suurpiiri')

    def test_parse_sap_range_row_2814E(self):
        """Test parsing a 2814E range from SAP format"""
        row = [
            MagicMock(value="Kalasatama"),
            MagicMock(value="2814E00001-2814E00999"),
        ]

        result = self.command._parse_sap_range_row(row)

        self.assertIsNotNone(result)
        self.assertEqual(result['projectTypePrefix'], '2814E')
        self.assertEqual(result['rangeStart'], '2814E00001')
        self.assertEqual(result['rangeEnd'], '2814E00999')
        self.assertEqual(result['area'], 'Kalasatama')
        self.assertIsNone(result['majorDistrict'])

    def test_parse_sap_range_row_invalid_format(self):
        """Test that invalid format returns None"""
        row = [
            MagicMock(value="Some District"),
            MagicMock(value="not-a-valid-range"),
        ]

        result = self.command._parse_sap_range_row(row)
        self.assertIsNone(result)

    def test_parse_sap_range_row_empty_range(self):
        """Test that empty range value returns None"""
        row = [
            MagicMock(value="ETELÄINEN SUURPIIRI"),
            MagicMock(value=None),
        ]

        result = self.command._parse_sap_range_row(row)
        self.assertIsNone(result)

    def test_parse_sap_range_row_no_dash_in_range(self):
        """Test that range without dash returns None"""
        row = [
            MagicMock(value="ETELÄINEN SUURPIIRI"),
            MagicMock(value="2814I00003"),
        ]

        result = self.command._parse_sap_range_row(row)
        self.assertIsNone(result)

    # =========================================================================
    # Asset Class Parsing Tests
    # =========================================================================

    def test_parse_asset_class_row_valid(self):
        """Test parsing a valid asset class row"""
        # Mock row: [category, name, componentClass, account, holdingPeriod, ...]
        row = [
            MagicMock(value="Aineelliset hyödykkeet"),  # category
            MagicMock(value="Maa- ja vesialueet"),      # name
            MagicMock(value="8103000"),                 # componentClass
            MagicMock(value="103000"),                  # account
            MagicMock(value="EP"),                      # holdingPeriod (no depreciation)
        ]

        result = self.command._parse_asset_class_row(row)

        self.assertIsNotNone(result)
        self.assertEqual(result['componentClass'], '8103000')
        self.assertEqual(result['account'], '103000')
        self.assertEqual(result['name'], 'Maa- ja vesialueet')
        self.assertFalse(result['hasHoldingPeriod'])  # EP = no depreciation
        self.assertIsNone(result['holdingPeriodYears'])
        self.assertEqual(result['category'], 'Aineelliset hyödykkeet')

    def test_parse_asset_class_row_with_holding_period(self):
        """Test parsing with numeric holding period"""
        row = [
            MagicMock(value="Rakennukset"),
            MagicMock(value="Rakennukset"),
            MagicMock(value="8110000"),
            MagicMock(value="110000"),
            MagicMock(value="50"),  # 50 years holding period
        ]

        result = self.command._parse_asset_class_row(row)

        self.assertIsNotNone(result)
        self.assertTrue(result['hasHoldingPeriod'])
        self.assertEqual(result['holdingPeriodYears'], 50)

    def test_parse_asset_class_row_skips_header(self):
        """Test that header rows are skipped"""
        row = [
            MagicMock(value="Poisto-ryhmä"),
            MagicMock(value="Käyttöomaisuusluokka"),
            MagicMock(value="Kom-luokka"),
            MagicMock(value="Tili"),
            MagicMock(value="Pito-aika"),
        ]

        result = self.command._parse_asset_class_row(row)
        self.assertIsNone(result)

    def test_parse_asset_class_row_missing_required_fields(self):
        """Test that missing required fields returns None"""
        row = [
            MagicMock(value="Category"),
            MagicMock(value=None),  # missing name
            MagicMock(value="8103000"),
            MagicMock(value="103000"),
            MagicMock(value="EP"),
        ]

        result = self.command._parse_asset_class_row(row)
        self.assertIsNone(result)

    # =========================================================================
    # District Pattern Matching Tests
    # =========================================================================


    def test_district_pattern_all_districts(self):
        """Test that all 8 major districts are recognized"""
        districts = [
            ("ETELÄINEN SUURPIIRI", "01", "Eteläinen suurpiiri"),
            ("LÄNTINEN SUURPIIRI", "02", "Läntinen suurpiiri"),
            ("KESKINEN SUURPIIRI", "03", "Keskinen suurpiiri"),
            ("POHJOINEN SUURPIIRI", "04", "Pohjoinen suurpiiri"),
            ("KOILLINEN SUURPIIRI", "05", "Koillinen suurpiiri"),
            ("KAAKKOINEN SUURPIIRI", "06", "Kaakkoinen suurpiiri"),
            ("ITÄINEN SUURPIIRI", "07", "Itäinen suurpiiri"),
            ("ÖSTERSUNDOM", "08", "Östersundomin suurpiiri"),
        ]

        for district_input, expected_code, expected_name in districts:
            row = [
                MagicMock(value=district_input),
                MagicMock(value="2814I00001-2814I00099"),
            ]
            result = self.command._parse_sap_range_row(row)

            self.assertIsNotNone(result, f"Failed for district: {district_input}")
            self.assertEqual(result['majorDistrict'], expected_code, f"Code mismatch for {district_input}")
            self.assertEqual(result['majorDistrictName'], expected_name, f"Name mismatch for {district_input}")


class TalpaImporterIntegrationTestCase(TestCase):
    """Integration tests for talpaimporter command using real (generated) Excel files"""

    def setUp(self):
        """Create a temporary Excel file"""
        self.temp_dir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.temp_dir, "test_talpa_import.xlsx")
        
        wb = openpyxl.Workbook()
        
        # 1. Asset Classes Sheet
        ws_assets = wb.active
        ws_assets.title = "Käyttöomaisuusluokat"
        ws_assets.append(["Category", "Name", "ComponentClass", "Account", "HoldingPeriod"]) # Header
        ws_assets.append(["Aineelliset", "Test Asset", "8103000", "103000", "EP"]) # Row 1

        # 2. Service Classes Sheet
        ws_services = wb.create_sheet("Palveluluokat")
        ws_services.append(["Header", "Code", "Name", "Desc"]) # Header
        ws_services.append(["", "4601", "Kadut ja yleiset", "Desc text"]) # Row 1
        
        # 3. 2814I Range Sheet
        ws_2814I = wb.create_sheet("2814I-projektinumerovälit")
        ws_2814I.append(["District info", "Range"]) # Header
        # Row: 2814I00003-2814I00300 (8 03 01 01) Katujen... ETELÄINEN SUURPIIRI
        ws_2814I.append(["2814I00003-2814I00300 (8 03 01 01) Desc, ETELÄINEN SUURPIIRI"]) 
        
        # 4. 2814E Range Sheet
        ws_2814E = wb.create_sheet("Esirakentaminen projektinumerov")
        ws_2814E.append(["Range info"]) # Header
        # Row: 2814E01000 - 2814E01500 (8 01 00) Desc
        ws_2814E.append(["2814E01000 - 2814E01500 (8 01 00) Desc Kalasatama"])

        wb.save(self.excel_path)

    def tearDown(self):
        """Remove temporary file"""
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def test_full_import_execution(self):
        """Test the full import process"""
        # Run command
        call_command('talpaimporter', file=self.excel_path)

        # Verify DB content
        
        # Ranges
        self.assertEqual(TalpaProjectNumberRange.objects.count(), 2)
        r1 = TalpaProjectNumberRange.objects.get(projectTypePrefix='2814I')
        self.assertEqual(r1.rangeStart, '2814I00003')
        self.assertEqual(r1.majorDistrict, '01') # Eteläinen
        
        r2 = TalpaProjectNumberRange.objects.get(projectTypePrefix='2814E')
        self.assertEqual(r2.rangeStart, '2814E01000')
        self.assertEqual(r2.area, 'Kalasatama')

        # Services
        # Note: The command also imports hardcoded defaults, so count might be higher than 1
        self.assertTrue(TalpaServiceClass.objects.filter(code='4601').exists())

        # Assets
        self.assertTrue(TalpaAssetClass.objects.filter(componentClass='8103000').exists())

    def test_dry_run_execution(self):
        """Test dry run does not create objects"""
        call_command('talpaimporter', file=self.excel_path, dry_run=True)

        self.assertEqual(TalpaProjectNumberRange.objects.count(), 0)

    def test_clear_existing_execution(self):
        """Test clear_existing flag"""
        # Create one dummy
        TalpaProjectNumberRange.objects.create(projectTypePrefix='2814I', rangeStart='X', rangeEnd='Y')
        self.assertEqual(TalpaProjectNumberRange.objects.count(), 1)

        # Run with clear-existing
        call_command('talpaimporter', file=self.excel_path, clear_existing=True)

        # Should have deleted dummy and imported 2 new ones
        self.assertEqual(TalpaProjectNumberRange.objects.count(), 2)
        self.assertFalse(TalpaProjectNumberRange.objects.filter(rangeStart='X').exists())

