from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class TalpaExcelService:
    """Generates Projektin avauslomake Infra Excel for Talpa submission."""

    COLUMNS = [
        ("A", "Talousarviokohdan numero", "_budgetAccountNumber", 25),
        ("B", "Talousarviokohdan nimi", "_budgetAccountName", 25),
        ("C", "Projektinumeroväli", "_numberRange", 25),
        ("D", "Malliprojekti", "templateProject", 15),
        ("E", "Laji", "_typeCode", 10),
        ("F", "Prioriteetti", "_typePriority", 12),
        ("G", "SAP nimi", "projectName", 25),
        ("H", "Osoite+postinumero", "_addressFull", 30),
        ("I", "Projekti alkaa -päättyy, huom.takuuaika", "_scheduleFull", 30),
        ("J", "Vastuuhenkilö", "responsiblePerson", 25),
        ("K", "Palveluluokka", "_serviceCode", 15),
        ("L", "Käyttöomaisuusluokat", "_assetComponent", 20),
        ("M", "Invest. profiili", "investmentProfile", 15),
        ("N", "Profiilin nimi", "profileName", 30),
        ("O", "Valmius", "readiness", 10),
    ]

    HEADER_STYLE = {
        "font": Font(bold=True, size=11),
        "fill": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

    CELL_STYLE = {
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }

    def _get_value(self, opening, field: str) -> str:
        if field.startswith("_"):
            return self._get_computed(opening, field)
        value = getattr(opening, field, None)
        return str(value) if value is not None else ""

    def _get_computed(self, opening, field: str) -> str:
        if field == "_budgetAccountNumber":
            # Extract number from "8 03 01 02 Perusparantaminen..."
            val = opening.budgetAccount or ""
            parts = val.split(" ")
            if len(parts) >= 4:
                return " ".join(parts[0:4])
            return val
        if field == "_budgetAccountName":
            # Extract name part after the number
            val = opening.budgetAccount or ""
            parts = val.split(" ")
            if len(parts) >= 5:
                return " ".join(parts[4:]).strip()
            return ""
        if field == "_numberRange" and opening.projectNumberRange:
            r = opening.projectNumberRange
            return f"{r.rangeStart} - {r.rangeEnd}"
        if field == "_typeCode" and opening.projectType:
            return opening.projectType.code or ""
        if field == "_typePriority" and opening.projectType:
            return opening.projectType.priority or ""
        if field == "_addressFull":
            addr = opening.streetAddress or ""
            zip_code = opening.postalCode or ""
            return f"{addr} {zip_code}".strip()
        if field == "_scheduleFull":
            start = opening.projectStartDate.strftime("%d.%m.%Y") if opening.projectStartDate else ""
            end = opening.projectEndDate.strftime("%d.%m.%Y") if opening.projectEndDate else ""
            if start or end:
                return f"{start} - {end}"
            return ""
        if field == "_serviceCode" and opening.serviceClass:
            return opening.serviceClass.code or ""
        if field == "_assetComponent" and opening.assetClass:
            return opening.assetClass.componentClass or ""
        return ""

    def generate_excel(self, opening) -> BytesIO:
        return self.generate_batch_excel([opening])

    def generate_batch_excel(self, openings: List) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Projektin avauslomake"

        for col, header, _, width in self.COLUMNS:
            cell = ws[f"{col}1"]
            cell.value = header
            cell.font = self.HEADER_STYLE["font"]
            cell.fill = self.HEADER_STYLE["fill"]
            cell.alignment = self.HEADER_STYLE["alignment"]
            cell.border = self.CELL_STYLE["border"]
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 30

        for row, opening in enumerate(openings, start=2):
            for col, _, field, _ in self.COLUMNS:
                cell = ws[f"{col}{row}"]
                cell.value = self._get_value(opening, field)
                cell.alignment = self.CELL_STYLE["alignment"]
                cell.border = self.CELL_STYLE["border"]

        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def get_filename(self, opening) -> str:
        if opening.project and opening.project.name:
            name = opening.project.name
            safe_name = "".join(c for c in name if c.isalnum() or c in " -_")[:50]
            return f"Talpa_avauslomake_{safe_name}.xlsx"
        return f"Talpa_avauslomake_{opening.id}.xlsx"