import datetime
from openpyxl import load_workbook
import re
from .hierarchy import buildHierarchiesAndProjects, getColor, MAIN_CLASS_COLOR
from ....services import PersonService, ProjectService, NoteService, ProjectPhaseService
from . import IExcelFileHandler
from django.contrib import messages
from io import BytesIO
import logging

logger = logging.getLogger("infraohjelmointi_api")


class PlanningFileHandler(IExcelFileHandler):
    """Excel file handler implementation for planning data"""

    def __init__(self) -> None:
        super().__init__()

        self.planningYear1 = None
        self.phaseColumnIndex = None

    def proceed_with_file(self, excel_path):
        logger.error(
            "\n\nReading project data from planning file {}\n".format(excel_path)
        )

        wb = load_workbook(excel_path, data_only=True, read_only=True)
        self.proceed_with_excel_data(wb=wb)

        logger.info("Planning file import done\n\n")

    def proceed_with_uploaded_file(self, request, uploadedFile):
        messages.info(
            request,
            "Reading project data from planning file {}".format(uploadedFile.name),
        )
        wb = load_workbook(
            filename=BytesIO(uploadedFile.read()), data_only=True, read_only=True
        )

        self.proceed_with_excel_data(wb=wb)
        messages.success(request, "Planning file import done\n\n")

    def proceed_with_excel_data(self, wb):
        skipables = [
            "none",
            "ylitysoikeus",
            "tae&tse raamit",
            "tae & tse raami",
            "tae&tse raami",
            "tae & tse raami",
            "ylityspaine",
            "ylitysoikeus yhteensä",
        ]

        main_class = [
            cel[0]
            for cel in wb.worksheets[0][2:11]
            if cel[0].value
            and re.match("^\d \d\d.+", cel[0].value.strip())
            and hex(int(getColor(wb, cel[0].fill.start_color), 16)) == MAIN_CLASS_COLOR
        ][0].value

        for sheetname in wb.sheetnames:
            logger.debug("\n\nHandling sheet {}\n".format(sheetname))

            sheet = wb[sheetname]
            rows = list(sheet.rows)

            for cellIndex, cell in enumerate(rows[1]):
                # looking for 1st the month in the row
                if str(cell.value) == "1":
                    # month is found, the field above should be year field
                    self.planningYear1 = (
                        int(rows[0][cellIndex].value)
                        if str(rows[0][cellIndex].value).isnumeric()
                        else None
                    )
                    break
            # Finding correct phase column
            for cellIndex, cell in enumerate(rows[1]):
                # looking for 1st the month in the row
                if "vaihe" in str(cell.value).lower():
                    self.phaseColumnIndex = cellIndex
                    break
            # self.planningYear1 = (
            #     int(rows[0][9].value) if str(rows[0][9].value).isnumeric() else None
            # )
            # self.constructionYear = (
            #     int(rows[0][21].value) if str(rows[0][21].value).isnumeric() else None
            # )
            buildHierarchiesAndProjects(
                wb=wb,
                rows=rows,
                skipables=skipables,
                main_class=main_class,
                project_handler=self.proceed_with_project_row,
            )

        logger.info("\n\nTotal rows handled  {}\n".format(len(rows)))

    def proceed_with_project_row(
        self, row, name, project_class, project_location, project_group
    ):
        sapNumber = row[1].value
        sapNetwork = row[2].value
        projectManager = row[4].value.strip() if row[4].value else None
        responsiblePerson = (
            PersonService.get_or_create_by_last_name(lastName=projectManager)[0]
            if projectManager and projectManager != "?"
            else None
        )
        pwNumber = str(row[27].value).strip().lower() if len(row) > 27 else None
        # get phases according to the correct column here
        excelPhases = str(row[self.phaseColumnIndex].value).strip().lower().split(" ")
        logger.debug("Project '{}' has PW id '{}'".format(name, pwNumber))

        try:
            project = ProjectService.get(
                name=name,
                projectClass=project_class,
                projectLocation=project_location,
                projectGroup=project_group,
            )
        except:
            project = ProjectService.create(
                name=name,
                projectClass=project_class,
                description="Kuvaus puuttuu",
                projectLocation=project_location,
                projectGroup=project_group,
            )
        # Check if excel file has any phases, set project phase to programming if true
        # Asked by Vesa
        if any(phase in excelPhases for phase in ["s", "m", "p", "k", "v", "t"]):
            project.phase = ProjectPhaseService.get_by_value(value="programming")
            project.programmed = True
        else:
            project.phase = ProjectPhaseService.get_by_value(value="proposal")
            project.programmed = False

        # Setting timeline dates according to the monthly cells in planning excel
        # First year planning, next 3 years construction by default
        # Asked by Vesa
        for cell in row[9:23]:
            # Check if this project is programmed for any month
            if str(cell.value) == "1":
                if self.planningYear1 != None:
                    project.planningStartYear = self.planningYear1
                    # First date of the first month given an year
                    project.estPlanningStart = datetime.datetime(
                        self.planningYear1, 1, 1
                    )
                    # Last date of last month given an year
                    project.estPlanningEnd = datetime.datetime(
                        self.planningYear1, 12, 31
                    )
                    project.constructionEndYear = self.planningYear1 + 3
                    # First date of First month given an year
                    project.estConstructionStart = datetime.datetime(
                        self.planningYear1 + 1, 1, 1
                    )
                    # Last date of last month given an year
                    project.estConstructionEnd = datetime.datetime(
                        self.planningYear1 + 3, 12, 31
                    )
                break

        project.sapProject = sapNumber
        project.sapNetwork = (
            [sapNetwork]
            if sapNetwork != None and not str(sapNetwork).strip() in ['"', "?"]
            else None
        )
        project.hkrId = (
            pwNumber
            if pwNumber != "none" and pwNumber != "?" and pwNumber != "x"
            else None
        )
        project.personPlanning = responsiblePerson
        project.save()

        notes = str((row[28].value if len(row) > 28 else "")).strip()
        if notes != "None" and notes != "?":
            NoteService.create(
                content=notes,
                project=project,
                byPerson=responsiblePerson
                if responsiblePerson
                else PersonService.get_or_create_by_name(
                    firstName="Excel", lastName="Integraatio"
                )[0],
            )
