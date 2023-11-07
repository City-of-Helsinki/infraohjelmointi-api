from openpyxl import load_workbook
import datetime

from .hierarchy import buildHierarchiesAndProjects, getColor, MAIN_CLASS_COLOR
from ....models.ProjectFinancial import ProjectFinancial
from ....services import (
    PersonService,
    ProjectService,
    NoteService,
    ProjectCategoryService,
    ProjectFinancialService,
    ProjectPhaseService,
)

from . import IExcelFileHandler
from django.contrib import messages
from io import BytesIO

import logging

logger = logging.getLogger("infraohjelmointi_api")


class BudgetFileHandler(IExcelFileHandler):
    """Excel file handler implementation for budget data"""

    def __init__(self) -> None:
        super().__init__()
        self.project_categories_map = {
            str(pc.value).lower(): pc for pc in ProjectCategoryService.list_all()
        }
        self.current_budget_year = datetime.date.today().year

    def proceed_with_file(self, excel_path):
        logger.info(
            "\nReading project data from budget file {}\n".format(excel_path)
        )

        wb = load_workbook(excel_path, data_only=True, read_only=True)

        skipables = [
            "none",
            "ylitysoikeus",
            "tae&tse raamit",
            "tae & tse raamit",
            "tae&tse raami",
            "tae & tse raami",
            "ylityspaine",
            "ylitysoikeus yhteensä",
        ]

        for sheetname in wb.sheetnames:
            logger.debug("\n\nHandling sheet {}\n".format(sheetname))

            sheet = wb[sheetname]
            rows = list(sheet.rows)
            year_heading = None
            # there is a need to find the year for the budget file
            for year in rows[8:11]:
                if str(year[19].value).lower() == "tae":
                    year_heading = "tae"
                elif year_heading == "tae" and str(year[19].value).isnumeric():
                    self.current_budget_year = year[19].value
                    break

            buildHierarchiesAndProjects(
                wb=wb,
                rows=rows,
                skipables=skipables,
                project_handler=self.proceed_with_project_row,
            )

        logger.info("\n\nTotal rows handled  {}\n".format(len(rows)))
        logger.info("Budget file import done\n\n")

    def proceed_with_uploaded_file(self, request, uploadedFile):
        messages.info(
            request,
            "Reading project data from uploaded budget file {}".format(
                uploadedFile.name
            ),
        )

        wb = load_workbook(
            filename=BytesIO(uploadedFile.read()), data_only=True, read_only=True
        )

        self.proceed_with_excel_data(wb=wb)

        messages.success(request, "Budget file import done\n\n")

    def proceed_with_excel_data(self, wb):
        skipables = [
            "none",
            "ylitysoikeus",
            "tae&tse raamit",
            "tae & tse raamit",
            "ylityspaine",
            "ylitysoikeus yhteensä",
        ]

        for sheetname in wb.sheetnames:
            logger.debug("\n\nHandling sheet {}\n".format(sheetname))

            sheet = wb[sheetname]
            rows = list(sheet.rows)
            year_heading = None
            # there is a need to find the year for the budget file
            for year in rows[8:11]:
                if str(year[19].value).lower() == "tae":
                    year_heading = "tae"
                elif year_heading == "tae" and str(year[19].value).isnumeric():
                    self.current_budget_year = year[19].value
                    break

            buildHierarchiesAndProjects(
                wb=wb,
                rows=rows,
                skipables=skipables,
                project_handler=self.proceed_with_project_row,
            )

        logger.info("\n\nTotal rows handled  {}\n".format(len(rows)))

    def proceed_with_project_row(
        self, row, name, project_class, project_location, project_group
    ):
        category = None
        try:
            category = self.project_categories_map[str(row[1].value).strip().lower()]
        except:
            # nothing to do here
            pass

        effectHousing = str(row[2].value).strip()
        effectHousing = True if effectHousing == "K" else False
        costForecast = row[6].value

        budgetProposalCurrentYearPlus0 = row[19].value
        budgetProposalCurrentYearPlus1 = row[20].value
        budgetProposalCurrentYearPlus2 = row[21].value
        preliminaryCurrentYearPlus3 = row[22].value
        preliminaryCurrentYearPlus4 = row[23].value
        preliminaryCurrentYearPlus5 = row[24].value
        preliminaryCurrentYearPlus6 = row[25].value
        preliminaryCurrentYearPlus7 = row[26].value
        preliminaryCurrentYearPlus8 = row[27].value
        preliminaryCurrentYearPlus9 = row[28].value

        budget_list = [
            0,
            budgetProposalCurrentYearPlus0 or 0,
            budgetProposalCurrentYearPlus1 or 0,
            budgetProposalCurrentYearPlus2 or 0,
            preliminaryCurrentYearPlus3 or 0,
            preliminaryCurrentYearPlus4 or 0,
            preliminaryCurrentYearPlus5 or 0,
            preliminaryCurrentYearPlus6 or 0,
            preliminaryCurrentYearPlus7 or 0,
            preliminaryCurrentYearPlus8 or 0,
            preliminaryCurrentYearPlus9 or 0,
        ]

        budget_sum = sum(budget_list)

        notes = str(row[29].value).strip()

        # some rows might not have pw number
        # with this we skip the index out of range -error
        
        pwNumber = None
        if not len(row) < 31:
            pwNumber = str(row[30].value).strip().lower()

        try:
            project = ProjectService.get(
                name=name,
                projectClass=project_class,
                projectLocation=project_location,
                projectGroup=project_group,
            )
        except:
            project = ProjectService.create(
                name=name,
                projectClass=project_class,
                description="Kuvaus puuttuu",
                projectLocation=project_location,
                projectGroup=project_group,
            )

        try:

            def find_non_zero_budget_index(budget_list):
                """
                Helper function to find first and last non zero budget
                """
                first_non_zero_index = None
                last_non_zero_index = None
                for i, num in enumerate(budget_list):
                    if num != 0:
                        if first_non_zero_index is None:
                            first_non_zero_index = i
                        last_non_zero_index = i

                return first_non_zero_index, last_non_zero_index

            project.phase = (
                ProjectPhaseService.get_by_value(value="programming")
                if budget_sum > 0
                else ProjectPhaseService.get_by_value(value="proposal")
            )

            firstBudgetIndex, lastBudgetIndex = find_non_zero_budget_index(
                budget_list=budget_list
            )

            if firstBudgetIndex != None and lastBudgetIndex != None:
                hasOneBudgetField = firstBudgetIndex == lastBudgetIndex
                planningStartYear = self.current_budget_year + firstBudgetIndex
                # set first budget year as planning start
                project.planningStartYear = planningStartYear
                # First date of the first month
                project.estPlanningStart = datetime.datetime(planningStartYear, 1, 1)
                project.frameEstPlanningStart = project.estPlanningStart
                # middle of year if only 1 budget in excel, else end of year
                project.estPlanningEnd = (
                    datetime.datetime(planningStartYear, 6, 30)
                    if hasOneBudgetField
                    else datetime.datetime(planningStartYear, 12, 31)
                )
                project.frameEstPlanningEnd = project.estPlanningEnd
                # same year as planning if 1 budget in excel, else the same year as last budget in excel
                project.constructionEndYear = (
                    (planningStartYear)
                    if hasOneBudgetField
                    else (self.current_budget_year + lastBudgetIndex)
                )

                # 1 month after planning ends if 1 budget field, else 1 year after planning
                project.estConstructionStart = (
                    datetime.datetime(planningStartYear, 7, 1)
                    if hasOneBudgetField
                    else datetime.datetime(planningStartYear + 1, 1, 1)
                )
                project.frameEstConstructionStart = project.estConstructionStart

                project.estConstructionEnd = datetime.datetime(
                    project.constructionEndYear, 12, 31
                )
                project.frameEstConstructionEnd = project.estConstructionEnd

            project.programmed = budget_sum > 0
            project.category = category
            project.effectHousing = effectHousing
            # if value already converted into float, convert it back to string to void validation error
            project.costForecast = str(costForecast) if costForecast != None else None
            project_financials = []
            for index, budgetValue in enumerate(budget_list):
                project_financials.append(
                    ProjectFinancial(
                        year=self.current_budget_year + index,
                        project_id=project.id,
                        value=str(budgetValue) if budgetValue != None else None,
                        forFrameView=False,
                    )
                )
                project_financials.append(
                    ProjectFinancial(
                        year=self.current_budget_year + index,
                        project_id=project.id,
                        value=str(budgetValue) if budgetValue != None else None,
                        forFrameView=True,
                    )
                )

            ProjectFinancialService.update_or_create_bulk(
                project_financials=project_financials
            )
            project.hkrId = (
                pwNumber
                if pwNumber != "none" and pwNumber != "?" and pwNumber != "x"
                else None
            )
            project.save()
        except Exception as e:
            print(f"Project {name} handling ended in exception")
            print(e)

        if notes != "None" and notes != "?":
            NoteService.create(
                content=notes,
                project=project,
                byPerson=None,
            )
