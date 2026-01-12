"""
Management command to import Talpa reference data from Excel file.

This imports:
- TalpaProjectNumberRange (from 2814I and 2814E project number range tabs)
- TalpaServiceClass (from Palveluluokat tab)
- TalpaAssetClass (from Käyttöomaisuusluokat tab)
- TalpaProjectType (from Talousarviokohdat and Lajit ja prioriteetit tabs)

Usage:
    python manage.py talpaimporter --file path/to/Projektin_avauslomake_Infra.xlsx
    python manage.py talpaimporter --file path/to/excel.xlsx --dry-run
"""

import os
import re
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
except ImportError:
    raise CommandError("openpyxl is required to read Excel files. Please install it.")

from infraohjelmointi_api.models import (
    TalpaProjectNumberRange,
    TalpaServiceClass,
    TalpaAssetClass,
    TalpaProjectType,
)


class Command(BaseCommand):
    help = """
    Import Talpa reference data from Excel file.
    
    This imports project number ranges, service classes, and asset classes
    from the official Talpa Excel specification file.
    
    Usage: python manage.py talpaimporter --file path/to/excel.xlsx
    """

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to the Talpa Excel file (Projektin avauslomake Infra)'
        )
        parser.add_argument(
            '--ranges-file',
            type=str,
            required=False,
            help='Path to SAP_Projektinumerovälit.xlsx for 2814I project number ranges'
        )
        parser.add_argument(
            '--preconstruction-file',
            type=str,
            required=False,
            help='Path to Ohjelmointityökalu_projektinumerovälit.xlsx for 2814E ranges with unit data'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Show what would be imported without making changes'
        )
        parser.add_argument(
            '--clear-existing',
            action='store_true',
            help='Clear all existing Talpa reference data before importing'
        )
        parser.add_argument(
            '--skip-ranges',
            action='store_true',
            help='Skip importing project number ranges'
        )
        parser.add_argument(
            '--skip-services',
            action='store_true',
            help='Skip importing service classes'
        )
        parser.add_argument(
            '--skip-assets',
            action='store_true',
            help='Skip importing asset classes'
        )
        parser.add_argument(
            '--skip-project-types',
            action='store_true',
            help='Skip importing project types and priorities'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        ranges_file_path = options.get('ranges_file')
        preconstruction_file_path = options.get('preconstruction_file')
        dry_run = options['dry_run']
        clear_existing = options['clear_existing']

        if not os.path.exists(file_path):
            raise CommandError(f"Excel file not found: {file_path}")

        self.stdout.write(self.style.NOTICE("Importing Talpa Reference Data from Excel..."))

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN MODE - No changes will be made\n"))

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as e:
            raise CommandError(f"Error reading Excel file: {e}")

        self.stdout.write(f"Opened: {file_path}")
        self.stdout.write(f"Worksheets found: {wb.sheetnames}\n")

        # Load ranges file if provided (for 2814I investment ranges)
        ranges_wb = None
        if ranges_file_path:
            if not os.path.exists(ranges_file_path):
                raise CommandError(f"Ranges Excel file not found: {ranges_file_path}")
            try:
                ranges_wb = openpyxl.load_workbook(ranges_file_path, data_only=True, read_only=True)
                self.stdout.write(f"Opened ranges file: {ranges_file_path}")
                self.stdout.write(f"Ranges worksheets: {ranges_wb.sheetnames}\n")
            except Exception as e:
                raise CommandError(f"Error reading ranges Excel file: {e}")

        # Load preconstruction file if provided (for 2814E with unit data)
        preconstruction_wb = None
        if preconstruction_file_path:
            if not os.path.exists(preconstruction_file_path):
                raise CommandError(f"Preconstruction Excel file not found: {preconstruction_file_path}")
            try:
                preconstruction_wb = openpyxl.load_workbook(preconstruction_file_path, data_only=True, read_only=True)
                self.stdout.write(f"Opened preconstruction file: {preconstruction_file_path}")
                self.stdout.write(f"Preconstruction worksheets: {preconstruction_wb.sheetnames}\n")
            except Exception as e:
                raise CommandError(f"Error reading preconstruction Excel file: {e}")

        stats = {
            'ranges_created': 0,
            'ranges_updated': 0,
            'services_created': 0,
            'services_updated': 0,
            'assets_created': 0,
            'assets_updated': 0,
            'priorities_updated': 0,
            'project_types_created': 0,
            'errors': []
        }

        # Pass workbooks to import functions
        options['_ranges_wb'] = ranges_wb
        options['_preconstruction_wb'] = preconstruction_wb

        if dry_run:
            self._dry_run_import(wb, options, stats)
        else:
            self._execute_import(wb, options, stats, clear_existing)

        self._print_summary(stats, dry_run)

    def _dry_run_import(self, wb, options, stats):
        """Preview what would be imported"""
        if not options['skip_assets']:
            self._preview_asset_classes(wb, stats)

        if not options['skip_services']:
            self._preview_service_classes(wb, stats)

        if not options['skip_ranges']:
            self._preview_project_number_ranges(wb, stats, options)

        if not options['skip_project_types']:
            self._preview_project_types(wb, stats)

    @transaction.atomic
    def _execute_import(self, wb, options, stats, clear_existing):
        """Execute the actual import"""
        if clear_existing:
            self._clear_existing_data(options, stats)

        if not options['skip_assets']:
            self._import_asset_classes(wb, stats)

        if not options['skip_services']:
            self._import_service_classes(wb, stats)

        if not options['skip_ranges']:
            self._import_project_number_ranges(wb, stats, options)

        if not options['skip_project_types']:
            self._import_project_types(wb, stats)

    def _clear_existing_data(self, options, stats):
        """Clear existing Talpa reference data"""
        self.stdout.write(self.style.WARNING("Clearing existing Talpa reference data..."))

        if not options['skip_ranges']:
            count = TalpaProjectNumberRange.objects.count()
            TalpaProjectNumberRange.objects.all().delete()
            self.stdout.write(f"Deleted {count} project number ranges")

        if not options['skip_services']:
            count = TalpaServiceClass.objects.count()
            TalpaServiceClass.objects.all().delete()
            self.stdout.write(f"Deleted {count} service classes")

        if not options['skip_assets']:
            count = TalpaAssetClass.objects.count()
            TalpaAssetClass.objects.all().delete()
            self.stdout.write(f"Deleted {count} asset classes")

        if not options['skip_project_types']:
             # We don't delete Project Types as they might be linked.
             pass

    # =========================================================================
    # ASSET CLASSES (Käyttöomaisuusluokat)
    # =========================================================================

    def _find_asset_class_sheet(self, wb):
        """Find the Käyttöomaisuusluokat worksheet"""
        for name in wb.sheetnames:
            if 'käyttöomaisuus' in name.lower() or 'kayttoomaisuus' in name.lower():
                return wb[name]
        return None

    def _parse_asset_class_row(self, row):
        """Parse a row from the asset class sheet"""
        try:
            component_class = str(row[2].value).strip() if row[2].value else None
            account = str(row[3].value).strip() if row[3].value else None
            holding_period_raw = str(row[4].value).strip() if row[4].value else None
            name = str(row[1].value).strip() if row[1].value else None

            if not component_class or not account or not name:
                return None

            # Skip header rows
            if component_class.lower() in ['kom-luokka', 'komponenttiluokka', 'tili']:
                return None

            # Parse holding period
            has_holding_period = True
            holding_period_years = None

            if holding_period_raw:
                if holding_period_raw.upper() == 'EP':
                    # EP = Ei poisteta (No depreciation)
                    has_holding_period = False
                else:
                    try:
                        holding_period_years = int(holding_period_raw)
                    except ValueError:
                        pass

            # Determine category from first column or component class prefix
            category = None
            if row[0].value:
                category = str(row[0].value).strip()

            return {
                'componentClass': component_class,
                'account': account,
                'name': name,
                'holdingPeriodYears': holding_period_years,
                'hasHoldingPeriod': has_holding_period,
                'category': category,
                'isActive': True,
            }
        except Exception:
            return None

    def _preview_asset_classes(self, wb, stats):
        """Preview asset class import"""
        self.stdout.write("Previewing Asset Classes...")

        sheet = self._find_asset_class_sheet(wb)
        if not sheet:
            self.stdout.write(self.style.WARNING("  Sheet not found"))
            return

        self.stdout.write(f"  Found sheet: {sheet.title}")
        count = 0
        for row in list(sheet.rows)[1:]:  # Skip header
            data = self._parse_asset_class_row(row)
            if data:
                count += 1
                if count <= 5:
                    self.stdout.write(f"  {count}. {data['componentClass']} - {data['name'][:40]}...")

        self.stdout.write(f"  Total asset classes found: {count}")
        stats['assets_created'] = count

    def _import_asset_classes(self, wb, stats):
        """Import asset classes"""
        self.stdout.write("Importing Asset Classes...")

        sheet = self._find_asset_class_sheet(wb)
        if not sheet:
            self.stdout.write(self.style.WARNING("  Sheet not found, skipping"))
            return

        current_category = None

        for row in list(sheet.rows)[1:]:
            # Check for category header in first column regardless of whether it's a data row
            col0 = str(row[0].value).strip() if row[0].value else None

            # Update 'current_category' if we see a value in Col 0
            # Ignore the specific table header 'Poisto-ryhmä'
            if col0 and 'poisto-ryhm' not in col0.lower():
                current_category = col0

            data = self._parse_asset_class_row(row)
            if data:
                # Apply the current category to the item
                if current_category:
                    data['category'] = current_category

                obj, created = TalpaAssetClass.objects.update_or_create(
                    componentClass=data['componentClass'],
                    account=data['account'],
                    defaults=data
                )
                if created:
                    stats['assets_created'] += 1
                else:
                    stats['assets_updated'] += 1

        self.stdout.write(f"  Created: {stats['assets_created']}, Updated: {stats['assets_updated']}")

    # =========================================================================
    # SERVICE CLASSES (Palveluluokat)
    # =========================================================================

    def _find_service_class_sheet(self, wb):
        """Find the Palveluluokat worksheet or section"""
        for name in wb.sheetnames:
            if 'palvelu' in name.lower():
                return wb[name]
        return None

    def _preview_service_classes(self, wb, stats):
        """Preview service class import"""
        self.stdout.write("Previewing Service Classes...")

        service_classes = [
            {'code': '4601', 'name': 'Kadut ja yleiset alueet', 'projectTypePrefix': '2814I'},
            {'code': '4701', 'name': 'Puistot ja viheralueet', 'projectTypePrefix': '2814I'},
            {'code': '3551', 'name': 'Liikunta- ja ulkoilupalvelut', 'projectTypePrefix': '2814I'},
            {'code': '5361', 'name': 'Maaomaisuuden hallinta', 'projectTypePrefix': '2814E'},
        ]

        sheet = self._find_service_class_sheet(wb)
        if sheet:
            self.stdout.write(f"  Found sheet: {sheet.title}")
        else:
            self.stdout.write("  No dedicated sheet found, using known values from form")

        for sc in service_classes:
            self.stdout.write(f"  {sc['code']} - {sc['name']} ({sc['projectTypePrefix']})")

        stats['services_created'] = len(service_classes)

    def _import_service_classes(self, wb, stats):
        """Import service classes"""
        self.stdout.write("Importing Service Classes...")

        service_classes = [
            {
                'code': '4601',
                'name': 'Kadut ja yleiset alueet',
                'description': 'Katujen, kevyen liikenteen väylien, jalkakäytävien ja liikennealueiden viheralueiden suunnittelu, toteutus, kunnossapito ja puhtaanapito',
                'projectTypePrefix': '2814I',
                'isActive': True,
            },
            {
                'code': '4701',
                'name': 'Puistot ja viheralueet',
                'description': 'Puistojen ja viheralueiden suunnittelu, toteutus, kunnossapito ja puhtaanapito',
                'projectTypePrefix': '2814I',
                'isActive': True,
            },
            {
                'code': '3551',
                'name': 'Liikunta- ja ulkoilupalvelut',
                'description': 'Yleisten edellytysten luominen liikunnalle paikallisella tasolla sekä kuntalaisten liikkumista tukevan toiminta, suunnittelu, toteutus sekä kunnossapito',
                'projectTypePrefix': '2814I',
                'isActive': True,
            },
            {
                'code': '5361',
                'name': 'Maaomaisuuden hallinta',
                'description': 'Maaomaisuuden hankinta, lunastaminen ja luovutus. Sisältää maaperätutkimukset, pilaantuneen maan puhdistamisen, kiinteistön kuntoarviot.',
                'projectTypePrefix': '2814E',
                'isActive': True,
            },
        ]

        for sc in service_classes:
            obj, created = TalpaServiceClass.objects.update_or_create(
                code=sc['code'],
                defaults=sc
            )
            if created:
                stats['services_created'] += 1
            else:
                stats['services_updated'] += 1

        self.stdout.write(f"  Created: {stats['services_created']}, Updated: {stats['services_updated']}")

    # =========================================================================
    # PROJECT NUMBER RANGES (2814I and 2814E)
    # =========================================================================

    def _find_project_range_sheets(self, wb):
        """Find the 2814I and 2814E project range worksheets"""
        sheets = {'2814I': None, '2814E': None}

        for name in wb.sheetnames:
            name_lower = name.lower()
            if '2814i' in name_lower:
                sheets['2814I'] = wb[name]
            elif 'esirakentaminen' in name_lower or '2814e' in name_lower:
                sheets['2814E'] = wb[name]

        return sheets

    def _parse_2814I_range_row(self, row):
        """Parse a row from the 2814I project number range sheet"""
        try:
            cell_value = str(row[0].value).strip() if row[0].value else None
            if not cell_value:
                return None

            if 'projektinumero' in cell_value.lower() or 'projektiväl' in cell_value.lower():
                return None

            cell_value = cell_value.replace('\u2013', '-').replace('\u2014', '-')

            # Regex: "2814I00003-2814I00300"
            range_pattern = r'^\s*(2814[IE]\d+)\s*[-]\s*(2814[IE]?\d+)'
            match = re.match(range_pattern, cell_value)

            if match:
                range_start = match.group(1)
                range_end = match.group(2)
            else:
                return None

            budget_account = None
            budget_account_number = None
            notes = None

            paren_match = re.search(r'\(([^)]+)\)', cell_value)
            if paren_match:
                paren_content = paren_match.group(1).strip()

                budget_match = re.match(r'^(8\s*\d{2}\s*\d{2}\s*\d{2}[A-H]?)', paren_content)

                if budget_match:
                    budget_account = budget_match.group(1).strip()
                    remaining = paren_content[budget_match.end():].strip()
                    if remaining.startswith(','):
                        notes = remaining[1:].strip()
                    elif remaining:
                        notes = remaining.strip()
                else:
                    if ',' in paren_content:
                        parts = paren_content.split(',', 1)
                        potential_budget = parts[0].strip()
                        if potential_budget.startswith('8') and len(potential_budget) <= 15:
                            budget_account = potential_budget
                            notes = parts[1].strip() if len(parts) > 1 else None
                        else:
                            notes = paren_content
                    else:
                        if paren_content.startswith('8') and len(paren_content) <= 15:
                            budget_account = paren_content
                        else:
                            notes = paren_content

                if budget_account:
                    budget_account_number = budget_account.replace(' ', '')

            district_name = None
            district_code = None

            district_patterns = [
                (r'ETELÄINEN\s+SUURPIIRI', '01', 'Eteläinen suurpiiri'),
                (r'LÄNTINEN\s+SUURPIIRI', '02', 'Läntinen suurpiiri'),
                (r'KESKINEN\s+SUURPIIRI', '03', 'Keskinen suurpiiri'),
                (r'POHJOINEN\s+SUURPIIRI', '04', 'Pohjoinen suurpiiri'),
                (r'KOILLINEN\s+SUURPIIRI', '05', 'Koillinen suurpiiri'),
                (r'KAAKKOINEN\s+SUURPIIRI', '06', 'Kaakkoinen suurpiiri'),
                (r'ITÄINEN\s+SUURPIIRI', '07', 'Itäinen suurpiiri'),
                (r'ÖSTERSUNDOM', '08', 'Östersundomin suurpiiri'),
            ]

            for pattern, code, name in district_patterns:
                if re.search(pattern, cell_value, re.IGNORECASE):
                    district_code = code
                    district_name = name
                    break

            return {
                'projectTypePrefix': '2814I',
                'budgetAccount': budget_account,
                'budgetAccountNumber': budget_account_number,
                'rangeStart': range_start,
                'rangeEnd': range_end,
                'majorDistrict': district_code,
                'majorDistrictName': district_name,
                'notes': notes,
                'isActive': True,
            }
        except Exception:
            return None

    def _parse_2814E_range_row(self, row):
        """Parse a row from the 2814E project number range sheet (legacy format from main file)"""
        try:
            cell_value = str(row[0].value).strip() if row[0].value else None
            if not cell_value:
                return None

            if 'projektinumero' in cell_value.lower() or '2814e-' in cell_value.lower():
                return None

            parts = cell_value.split()
            if len(parts) < 2:
                return None

            range_part = parts[0]
            range_start = range_part.strip()
            range_end = range_start

            for i, part in enumerate(parts):
                if part == '-' and i + 1 < len(parts):
                    range_end = parts[i + 1].strip()
                    break
                elif '-' in part and part != range_start:
                    range_split = part.split('-')
                    if len(range_split) == 2:
                        range_end = range_split[1].strip()
                    break

            budget_account = None
            budget_account_number = None
            budget_match = re.search(r'\(?(8\s*\d{2}\s*\d{2}\s*\d{2})\)?', cell_value)
            if budget_match:
                budget_account = budget_match.group(1).strip()
                budget_account_number = budget_account.replace(' ', '')

            area = None
            area_patterns = [
                'Kalasatama', 'Länsisatama', 'Pasila', 'Kruunuvuorenranta',
                'Kuninkaankolmio', 'Malmi', 'Malminkartano', 'Mellunkylä',
                'Meri-Rastila', 'Läntinen bulevardikaupunki', 'Makasiiniranta'
            ]
            for area_name in area_patterns:
                if area_name.lower() in cell_value.lower():
                    area = area_name
                    break

            return {
                'projectTypePrefix': '2814E',
                'budgetAccount': budget_account,
                'budgetAccountNumber': budget_account_number,
                'rangeStart': range_start,
                'rangeEnd': range_end,
                'area': area,
                'unit': None,  # Legacy format doesn't have unit column
                'isActive': True,
            }
        except Exception:
            return None

    def _parse_preconstruction_row(self, row, current_budget_account=None, current_area=None):
        """
        Parse a row from Ohjelmointityökalu_projektinumerovälit.xlsx format.
        
        Format: TA-kohta | Area | Yksikkö | RangeStart | RangeEnd | ContactPerson | | Email
        Example: 8 08 01 03|Kalasatama|Tontit|2814E22001|2814E22099|Satu Järvinen||Satu.Jarvinen@hel.fi
        
        Note: Some rows are continuations where TA-kohta and Area are empty but Unit/Range are filled.
        """
        try:
            # Column A (0): TA-kohta / Budget Account (may be empty for continuation rows)
            col_a = str(row[0].value).strip() if row[0].value else None
            # Column B (1): Area name (may be empty for continuation rows)
            col_b = str(row[1].value).strip() if len(row) > 1 and row[1].value else None
            # Column C (2): Yksikkö / Unit - "Tontit", "Mao", or "Geo"
            col_c = str(row[2].value).strip() if len(row) > 2 and row[2].value else None
            # Column D (3): Range Start - e.g., "2814E22001"
            col_d = str(row[3].value).strip() if len(row) > 3 and row[3].value else None
            # Column E (4): Range End - e.g., "2814E22099"
            col_e = str(row[4].value).strip() if len(row) > 4 and row[4].value else None

            # Skip header rows
            if col_a and ('ta-kohta' in col_a.lower() or 'make-palvelu' in col_a.lower()):
                return None, current_budget_account, current_area
            
            # Skip rows without valid range data
            if not col_d or not col_d.startswith('2814E'):
                return None, current_budget_account, current_area
            
            # Update current context if this row has TA-kohta or Area
            if col_a and col_a.startswith('8 '):
                current_budget_account = col_a
            if col_b and col_b not in ['', 'None']:
                current_area = col_b
            
            # Validate unit
            valid_units = ['Tontit', 'Mao', 'Geo']
            unit = col_c if col_c in valid_units else None
            
            return {
                'projectTypePrefix': '2814E',
                'budgetAccount': current_budget_account,
                'budgetAccountNumber': current_budget_account.replace(' ', '') if current_budget_account else None,
                'rangeStart': col_d,
                'rangeEnd': col_e or col_d,  # Use start as end if end is missing
                'area': current_area,
                'unit': unit,
                'isActive': True,
            }, current_budget_account, current_area
        except Exception:
            return None, current_budget_account, current_area

    def _parse_sap_range_row(self, row):
        """Parse a row from SAP_Projektinumerovälit.xlsx format"""
        try:
            area_or_district = str(row[0].value).strip() if row[0].value else None
            range_value = str(row[1].value).strip() if row[1].value else None

            if not range_value:
                return None

            if '-' not in range_value:
                return None

            range_parts = range_value.split('-')
            range_start = range_parts[0].strip()
            range_end = range_parts[1].strip() if len(range_parts) > 1 else range_start

            if not range_start.startswith('2814'):
                return None

            if range_start.startswith('2814I') or 'I' in range_start[4:5].upper():
                project_type_prefix = '2814I'
            elif range_start.startswith('2814E') or 'E' in range_start[4:5].upper():
                project_type_prefix = '2814E'
            else:
                project_type_prefix = '2814I'

            district_code = None
            district_name = None
            area = None

            if area_or_district:
                district_patterns = [
                    (r'ETELÄINEN', '01', 'Eteläinen suurpiiri'),
                    (r'LÄNTINEN', '02', 'Läntinen suurpiiri'),
                    (r'KESKINEN', '03', 'Keskinen suurpiiri'),
                    (r'POHJOINEN', '04', 'Pohjoinen suurpiiri'),
                    (r'KOILLINEN', '05', 'Koillinen suurpiiri'),
                    (r'KAAKKOINEN', '06', 'Kaakkoinen suurpiiri'),
                    (r'ITÄINEN', '07', 'Itäinen suurpiiri'),
                    (r'ÖSTERSUNDOM', '08', 'Östersundomin suurpiiri'),
                ]

                for pattern, code, name in district_patterns:
                    if re.search(pattern, area_or_district, re.IGNORECASE):
                        district_code = code
                        district_name = name
                        break

                if not district_name:
                    area = area_or_district

            return {
                'projectTypePrefix': project_type_prefix,
                'budgetAccount': None,
                'budgetAccountNumber': None,
                'rangeStart': range_start,
                'rangeEnd': range_end,
                'majorDistrict': district_code,
                'majorDistrictName': district_name,
                'area': area if project_type_prefix == '2814E' else None,
                'isActive': True,
            }
        except Exception:
            return None

    def _preview_project_number_ranges(self, wb, stats, options=None):
        """Preview project number range import"""
        self.stdout.write("Previewing Project Number Ranges...")

        ranges_wb = options.get('_ranges_wb') if options else None

        if ranges_wb:
            self.stdout.write("  Using separate ranges file (SAP format)")
            total = 0
            for sheet_name in ranges_wb.sheetnames:
                sheet = ranges_wb[sheet_name]
                self.stdout.write(f"\n  Processing sheet: {sheet_name}")
                count = 0
                for row in list(sheet.rows)[1:]:  # Skip header
                    data = self._parse_sap_range_row(row)
                    if data:
                        count += 1
                        if count <= 5:
                            self.stdout.write(
                                f"    {data['rangeStart']} - {data['rangeEnd']} "
                                f"({data.get('majorDistrictName') or data.get('area') or 'No location'})"
                            )

                self.stdout.write(f"  Total ranges from {sheet_name}: {count}")
                total += count

            stats['ranges_created'] = total
            return

        sheets = self._find_project_range_sheets(wb)

        total = 0
        for prefix, sheet in sheets.items():
            if sheet:
                self.stdout.write(f"\n  {prefix} sheet: {sheet.title}")
                count = 0
                for row in list(sheet.rows)[1:]:
                    if prefix == '2814I':
                        data = self._parse_2814I_range_row(row)
                    else:
                        data = self._parse_2814E_range_row(row)

                    if data:
                        count += 1
                        if count <= 3:
                            self.stdout.write(f"    {data['rangeStart']} - {data['rangeEnd']}")

                self.stdout.write(f"  Total {prefix} ranges: {count}")
                total += count
            else:
                self.stdout.write(f"\n  {prefix} sheet: NOT FOUND")

        stats['ranges_created'] = total

    def _import_project_number_ranges(self, wb, stats, options=None):
        """Import project number ranges"""
        self.stdout.write("Importing Project Number Ranges...")

        ranges_wb = options.get('_ranges_wb') if options else None
        preconstruction_wb = options.get('_preconstruction_wb') if options else None

        # Import from SAP ranges file (2814I)
        if ranges_wb:
            self.stdout.write("  Using separate ranges file (SAP format) for 2814I...")
            for sheet_name in ranges_wb.sheetnames:
                sheet = ranges_wb[sheet_name]
                for row in list(sheet.rows)[1:]:
                    data = self._parse_sap_range_row(row)
                    if data:
                        obj, created = TalpaProjectNumberRange.objects.update_or_create(
                            projectTypePrefix=data['projectTypePrefix'],
                            rangeStart=data['rangeStart'],
                            rangeEnd=data['rangeEnd'],
                            defaults=data
                        )
                        if created:
                            stats['ranges_created'] += 1
                        else:
                            stats['ranges_updated'] += 1

        # Import from preconstruction file (2814E with proper unit values)
        if preconstruction_wb:
            self.stdout.write("  Using Ohjelmointityökalu file for 2814E with unit values...")
            for sheet_name in preconstruction_wb.sheetnames:
                sheet = preconstruction_wb[sheet_name]
                current_budget_account = None
                current_area = None
                for row in list(sheet.rows)[1:]:  # Skip header row
                    result = self._parse_preconstruction_row(row, current_budget_account, current_area)
                    if result[0]:  # result is (data, budget_account, area)
                        data, current_budget_account, current_area = result
                        obj, created = TalpaProjectNumberRange.objects.update_or_create(
                            projectTypePrefix=data['projectTypePrefix'],
                            rangeStart=data['rangeStart'],
                            rangeEnd=data['rangeEnd'],
                            defaults=data
                        )
                        if created:
                            stats['ranges_created'] += 1
                        else:
                            stats['ranges_updated'] += 1
                    else:
                        # Update context even if no data returned
                        _, current_budget_account, current_area = result

        # Fallback: Import from main file sheets if no separate files provided
        if not ranges_wb and not preconstruction_wb:
            sheets = self._find_project_range_sheets(wb)

            for prefix, sheet in sheets.items():
                if not sheet:
                    continue

                for row in list(sheet.rows)[1:]:
                    if prefix == '2814I':
                        data = self._parse_2814I_range_row(row)
                    else:
                        data = self._parse_2814E_range_row(row)

                    if data:
                        obj, created = TalpaProjectNumberRange.objects.update_or_create(
                            projectTypePrefix=data['projectTypePrefix'],
                            rangeStart=data['rangeStart'],
                            rangeEnd=data['rangeEnd'],
                            defaults=data
                        )
                        if created:
                            stats['ranges_created'] += 1
                        else:
                            stats['ranges_updated'] += 1

        self.stdout.write(f"  Created: {stats['ranges_created']}, Updated: {stats['ranges_updated']}")

    # =========================================================================
    # PROJECT TYPES (Talousarviokohdat + Lajit ja prioriteetit)
    # =========================================================================

    def _find_sheet_by_keyword(self, wb, keyword):
        """Find sheet containing keyword in title"""
        for name in wb.sheetnames:
            if keyword.lower() in name.lower():
                return wb[name]
        return None

    def _parse_laji_header_cell(self, cell_value):
        """
        Parse Laji ID, Name, and Category from header cell.
        Returns (laji_code, search_name, category) or (None, None, None).

        Format: "01 (Uudisrakentaminen, KADUT, LIIKENNEVÄYLÄT JA RADAT)"
        - First comma-separated value = Laji Name
        - Second comma-separated value = Category
        """
        laji_code = None
        search_name = "Unknown"
        category = None

        # Try Standard format: NN (Name, Category...) or NN (Name)
        # Replaced regex with string splitting to avoid ReDoS (python:S5852)
        # Format: "01 (Uudisrakentaminen, KADUT, ...)"

        # 1. Extract ID (digits at start)
        first_space = cell_value.find(' ')
        if first_space > 0 and cell_value[:first_space].isdigit():
            laji_code = cell_value[:first_space]

            # 2. Check for parentheses content
            start_paren = cell_value.find('(')
            end_paren = cell_value.rfind(')')

            if start_paren > first_space and end_paren > start_paren:
                content = cell_value[start_paren+1:end_paren]

                # 3. Split content by comma
                # Use maxsplit=1 because Category may contain punctuation (e.g. "KADUT, RADAT")
                parts = [p.strip() for p in content.split(',', 1)]
                if parts:
                    search_name = parts[0]
                    if len(parts) > 1:
                        category = parts[1].title()

            return laji_code, search_name, category

        # Try Irregular "H (Laji X, Name...)" format
        if 'laji' in cell_value.lower():
            # "H (Laji 13, Uudet puistot, PUISTOT)"
            # Find "Laji NN" first
            laji_start = cell_value.lower().find('laji')
            if laji_start != -1:
                # Extract part after "Laji "
                substring = cell_value[laji_start+5:] # Skip "Laji "

                # Find space or comma ending the ID
                # "13, Uudet puistot..."
                separator = -1
                for i, char in enumerate(substring):
                    if not char.isdigit():
                        separator = i
                        break

                if separator > 0:
                    laji_code = substring[:separator]

                    # Parse the rest
                    # Expected: ", Uudet puistot, PUISTOT)"
                    rest = substring[separator:]

                    # Remove closing paren if present
                    end_paren = rest.rfind(')')
                    if end_paren != -1:
                        rest = rest[:end_paren]

                    parts = [p.strip() for p in rest.split(',')]
                    # parts[0] might be empty if separator was comma
                    real_parts = [p for p in parts if p]

                    if real_parts:
                        search_name = real_parts[0]
                        if len(real_parts) > 1:
                            category = real_parts[1].title()

                    return laji_code, search_name, category

        return None, None, None

    def _collect_priorities(self, sheet):
        """
        Parse Lajit sheet to collect metadata for each Laji group.
        Returns: { 'LajiCode': { 'priorities': [], 'category': '', 'search_name': '' } }

        Category is extracted from within the Laji header cell itself,
        not from separate rows.
        """
        laji_data = {}
        current_laji_code = None

        for row in sheet.rows:
            col0 = str(row[0].value).strip() if row[0].value else None
            col1 = str(row[1].value).strip() if row[1].value and len(str(row[1].value)) < 10 else None
            col3 = str(row[3].value).strip() if len(row) > 3 and row[3].value else None

            if col0:
                # Attempt to parse Laji Header (includes embedded category)
                laji_code, search_name, category = self._parse_laji_header_cell(col0)
                if laji_code:
                    # Fallback Category Logic for Inconsistent Excel Data
                    if not category or category == 'None' or len(category) < 3:
                        if laji_code in ['06', '07', '08']:
                            category = "Projektialueiden Kadut"
                        elif laji_code in ['15', '16', '17', '18', '19']:
                            category = "Projektialueiden Puistot"
                        elif laji_code == '13':
                            category = "Puistorakentaminen"
                        elif laji_code in ['01', '02', '03', '04', '05']:
                             category = "Kadut, Liikenneväylät Ja Radat"
                        elif laji_code in ['09', '10', '11', '12', '14']:
                             category = "Puistot Ja Liikunta-Alueet"
                        elif laji_code in ['20', '21']:
                             category = "Esirakentaminen"
                        elif laji_code in ['22', '23', '24']:
                             category = "Muut Yleiset Alueet"
                        elif laji_code == '25':
                             category = "Maaomaisuuden Hallinta"
                        elif laji_code in ['26', '27', '28', '33']:
                             category = "Kaupunkiuudistusalueet"
                        elif laji_code in ['29', '30', '31', '32']:
                             category = "Suuret Liikennehankkeet"
                        elif laji_code in ['34', '35', '36']:
                             category = "Projektialueiden Kadut"
                        elif laji_code in ['37', '38', '39']:
                             category = "Projektialueiden Puistot"

                if laji_code:
                    current_laji_code = laji_code
                    # Prevent overwriting if Laji code appears multiple times (use first occurrence)
                    if current_laji_code not in laji_data:
                        # Clean Name if it has prefixes like "8 09 01 "
                        # e.g. "8 09 01 Malminkartano" -> "Malminkartano"
                        search_name = re.sub(r'^[\d\s]{3,}', '', search_name).strip()
                        if search_name.isupper():
                             search_name = search_name.capitalize()

                        laji_data[current_laji_code] = {
                            'priorities': [],
                            'category': category,
                            'search_name': search_name
                        }
                    continue

            # Check for Priority Row (col1 = 'A', 'B', etc.)
            val1 = str(row[1].value).strip() if len(row) > 1 and row[1].value else None
            # Priority rows usually have a single letter or number in Col 1
            if val1 and (len(val1) <= 2 or val1.isdigit()):
                if current_laji_code and current_laji_code in laji_data:
                    prio = val1
                    desc = None
                    # Search for description in columns C, D, E (Indexes 2, 3, 4)
                    # Some rows have description in Col 2, others in Col 3
                    for col_idx in range(2, min(5, len(row))):
                        val = str(row[col_idx].value).strip() if row[col_idx].value else None
                        if val:
                            desc = val
                            break

                    if desc:
                        # Clean Description "A (Laji 01, Real Name)" -> "Real Name"
                        # Replaced regex with string logic
                        if '(laji' in desc.lower():
                            laji_idx = desc.lower().find('(laji')
                            if laji_idx != -1:
                                # Extract content inside parens of "(Laji ...)"
                                end_paren = desc.rfind(')', laji_idx)
                                if end_paren > laji_idx:
                                    inner = desc[laji_idx+1:end_paren] # "Laji 01, Real Name"

                                    # Split by comma
                                    parts = inner.split(',', 1)
                                    if len(parts) > 1:
                                        desc = parts[1].strip()

                        laji_data[current_laji_code]['priorities'].append({
                            'priority': prio,
                            'description': desc
                        })

        return laji_data



    def _format_name_with_id(self, original_name, laji_id):
        """Prepend Laji ID to name if not already present."""
        if laji_id and not original_name.startswith(laji_id):
             return f"{laji_id} {original_name}"
        return original_name

    def _preview_project_types(self, wb, stats):
        """Preview Project Types & Priorities import"""
        self.stdout.write("Previewing Project Types & Priorities...")

        lajit_sheet = self._find_sheet_by_keyword(wb, 'lajit')

        if not lajit_sheet:
            self.stdout.write(self.style.WARNING("  Required 'Lajit' sheet not found"))
            return

        laji_map = self._collect_priorities(lajit_sheet)
        self.stdout.write(f"  Found {len(laji_map)} Priority Groups (Lajit)")

        count = 0
        for laji_id, info in laji_map.items():
            count += 1
            if count <= 5:
                prios = info.get('priorities', [])
                self.stdout.write(f"  {laji_id} {info['search_name']} ({len(prios)} priorities)")
                if prios:
                    self.stdout.write(f"    - First Prio: {prios[0]['description']}")

        stats['project_types_created'] = 0

    def _import_project_types(self, wb, stats):
        """Import Project Types and apply Priorities"""
        self.stdout.write("Importing Project Types & Priorities...")

        lajit_sheet = self._find_sheet_by_keyword(wb, 'lajit')

        if not lajit_sheet:
            self.stdout.write(self.style.WARNING("  Required 'Lajit' sheet not found, skipping"))
            return

        # 1. Collect Metadata
        laji_map = self._collect_priorities(lajit_sheet)
        self.stdout.write(f"  Loaded metadata for {len(laji_map)} Laji groups")

        if not laji_map:
            self.stdout.write("No Laji data collected!")
            return

        created_count = 0

        # 3. Create Project Types from Collected Laji Data
        # We iterate the Clean Laji Data (Source of Truth) instead of messy Budget Lines
        for laji_id, info in laji_map.items():
            laji_name = info['search_name'] # e.g. "Uudisrakentaminen"
            category = info['category']     # e.g. "Kadut, liikenneväylät ja radat"
            priorities = info['priorities'] # List of dicts

            # 1. Create Base Laji Item (e.g. "01 Uudisrakentaminen")
            base_name = self._format_name_with_id(laji_name, laji_id)

            TalpaProjectType.objects.update_or_create(
                code=laji_id,
                priority=None,
                defaults={
                    'name': base_name,
                    'isActive': True,
                    'category': category,
                    # Ensure numerical sorting (1, 2, ... 10)
                    'sortOrder': int(laji_id) if laji_id.isdigit() else 0
                }
            )
            created_count += 1

            # 2. Create Priority Items
            for i, p in enumerate(priorities):
                p_code = p['priority']
                p_desc = p['description']

                # Handling missing priority codes to avoid collision with Base Item (priority=None)
                if not p_code:
                     p_code = f"SUB_{i+1}"

                if p_desc:
                    p_name = self._format_name_with_id(p_desc, laji_id)

                    TalpaProjectType.objects.update_or_create(
                        code=laji_id, # Same Laji ID
                        priority=p_code, # Unique per Laji
                        defaults={
                            'name': p_name,
                            'isActive': True,
                            'category': category,
                            'sortOrder': int(laji_id) if laji_id.isdigit() else 0
                        }
                    )
                    created_count += 1

        self.stdout.write(f"  Processed {created_count} project type definitions")
        stats['project_types_created'] = created_count

    def _print_summary(self, stats, dry_run):
        """Print import summary"""
        self.stdout.write(self.style.SUCCESS("IMPORT SUMMARY"))

        self.stdout.write(f"Asset Classes:   {stats['assets_created']} created, {stats['assets_updated']} updated")
        self.stdout.write(f"Service Classes: {stats['services_created']} created, {stats['services_updated']} updated")
        self.stdout.write(f"Number Ranges:   {stats['ranges_created']} created, {stats['ranges_updated']} updated")
        self.stdout.write(f"Project Types:   {stats.get('project_types_created', 0)} created/updated")

        if stats['errors']:
            self.stdout.write(self.style.ERROR(f"\nErrors ({len(stats['errors'])}):"))
            for error in stats['errors'][:10]:
                self.stdout.write(f"  - {error}")
            if len(stats['errors']) > 10:
                self.stdout.write(f"  ... and {len(stats['errors']) - 10} more")

        if dry_run:
            self.stdout.write(self.style.WARNING("\nThis was a dry run. No changes were made."))
        else:
            self.stdout.write(self.style.SUCCESS("\nImport completed successfully!"))
