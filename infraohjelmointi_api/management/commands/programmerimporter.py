"""
Management command to import programmers from Excel file
and assign them as default programmers to their designated project classes.
"""

import os
import uuid
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction, IntegrityError
from django.utils import timezone

try:
    import openpyxl
except ImportError:
    raise CommandError("openpyxl is required to read Excel files. Please install it.")

from infraohjelmointi_api.utils.project_class_utils import get_programmer_from_hierarchy
from infraohjelmointi_api.models import (
    ProjectClass,
    ProjectProgrammer,
    Person,
    Project
)


class Command(BaseCommand):
    help = """
    Import programmers from Excel file and assign them as default programmers
    to their designated project classes. Handles both specific assignments and fallback assignments.
    
    Usage: python manage.py programmerimporter --file path/to/excel
    """

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to the Excel file containing programmer assignments'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Show what would be imported without making changes'
        )
        parser.add_argument(
            '--clear-existing',
            action='store_true',
            help='Clear all existing defaultProgrammer assignments first'
        )
        parser.add_argument(
            '--skip-projects',
            action='store_true',
            help='Skip applying default programmers to existing projects (only set class defaults)'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        dry_run = options['dry_run']
        clear_existing = options['clear_existing']
        skip_projects = options['skip_projects']
        apply_to_projects = not skip_projects  # Apply by default unless skipped

        if not os.path.exists(file_path):
            raise CommandError(f"Excel file path is incorrect or missing: {file_path}")

        self.stdout.write(
            self.style.NOTICE(
                "\n" + '\033[94m' +
                "----------------------------------------------------------------\n" +
                "                 Importing Programmers from Excel               \n" +
                "----------------------------------------------------------------\n" +
                '\033[0m'
            )
        )

        self.stdout.write(f"Reading programmers from: {file_path}")

        try:
            specific_assignments, fallback_assignments = self.read_excel_file(file_path)
        except Exception as e:
            raise CommandError(f"Error reading Excel file: {e}")

        if not specific_assignments and not fallback_assignments:
            self.stdout.write(self.style.WARNING("No programmer assignments found in Excel file"))
            return

        total_assignments = len(specific_assignments) + len(fallback_assignments)
        self.stdout.write(f"Found {len(specific_assignments)} specific assignments")
        self.stdout.write(f"Found {len(fallback_assignments)} fallback assignments")
        self.stdout.write(f"Total: {total_assignments} assignments")

        if dry_run:
            self.show_dry_run(specific_assignments, fallback_assignments, clear_existing, apply_to_projects)
        else:
            self.process_assignments(specific_assignments, fallback_assignments, clear_existing, apply_to_projects)

    def read_excel_file(self, file_path):
        """Read the Excel file and extract programmer assignments"""
        specific_assignments = []
        fallback_assignments = []

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        self.stdout.write(f"Reading worksheet: {ws.title}")
        self.stdout.write(f"Total rows: {ws.max_row}, columns: {ws.max_column}")

        # Find the fallback section starting row
        fallback_start_row = None
        for row_num in range(1, ws.max_row + 1):
            col1 = ws.cell(row=row_num, column=1).value
            if col1 and 'Muut kohdat lokaatiotiedon mukaan' in str(col1):
                fallback_start_row = row_num + 1  # Start from next row
                self.stdout.write(f"Found fallback section starting at row {fallback_start_row}")
                break

        # Process specific assignments (before fallback section)
        end_row = fallback_start_row - 2 if fallback_start_row else ws.max_row
        for row_num in range(2, end_row + 1):  # Start from row 2 (skip header)
            assignment = self.parse_assignment_row(ws, row_num, is_fallback=False)
            if assignment:
                specific_assignments.append(assignment)

        # Process fallback assignments (after fallback section header)
        if fallback_start_row:
            for row_num in range(fallback_start_row, ws.max_row + 1):
                assignment = self.parse_assignment_row(ws, row_num, is_fallback=True)
                if assignment:
                    fallback_assignments.append(assignment)

        return specific_assignments, fallback_assignments

    def parse_assignment_row(self, worksheet, row_num, is_fallback=False):
        """Parse a single row into an assignment"""
        if is_fallback:
            # Fallback format: District name in col1, empty col2, programmer in col3
            district_name = worksheet.cell(row=row_num, column=1).value
            programmer_name = worksheet.cell(row=row_num, column=3).value

            if not district_name or not programmer_name:
                return None

            district_name = str(district_name).strip()
            programmer_name = str(programmer_name).strip()

            # Skip entries marked as empty
            if programmer_name.lower() in ['jätetään tyhjiksi', 'ei valintaa', '']:
                return None

            return {
                'type': 'fallback',
                'district_name': district_name,
                'programmer_name': programmer_name,
                'row': row_num
            }
        else:
            # Specific format: Class code in col1, description in col2, programmer in col3
            class_code = worksheet.cell(row=row_num, column=1).value
            class_description = worksheet.cell(row=row_num, column=2).value
            programmer_name = worksheet.cell(row=row_num, column=3).value

            if not class_code or not programmer_name:
                return None

            class_code = str(class_code).strip()
            programmer_name = str(programmer_name).strip()

            # Skip entries marked as empty
            if programmer_name.lower() in ['jätetään tyhjiksi', 'ei valintaa', '']:
                return None

            return {
                'type': 'specific',
                'class_code': class_code,
                'class_description': str(class_description).strip() if class_description else '',
                'programmer_name': programmer_name,
                'row': row_num
            }

    def parse_programmer_name(self, programmer_name):
        """Parse programmer name into first and last name"""
        parts = programmer_name.strip().split()
        if len(parts) >= 2:
            first_name = parts[0]
            last_name = ' '.join(parts[1:])
        else:
            first_name = programmer_name
            last_name = ''

        return first_name, last_name

    def find_project_class(self, class_code, class_description):
        """Find project class by code and description"""
        # Try to find exact match first
        classes = ProjectClass.objects.filter(name__icontains=class_code)

        if not classes.exists():
            return None, f"No project class found containing code '{class_code}'"

        if classes.count() == 1:
            return classes.first(), None

        # If multiple matches, try to find the best one using description
        if class_description:
            for cls in classes:
                if class_description.lower() in cls.name.lower():
                    return cls, None

        # Return the first match if no exact description match
        return classes.first(), f"Multiple classes found for '{class_code}', using first match"

    def find_classes_for_district_fallback(self, district_name):
        """Find all project classes that should use this district fallback"""
        # Find classes that contain the district name and don't have a specific programmer assigned
        classes = ProjectClass.objects.filter(
            name__icontains=district_name,
            defaultProgrammer__isnull=True
        )

        return classes

    def find_or_create_programmer(self, programmer_name, dry_run=False):
        """Find or create a ProjectProgrammer"""
        first_name, last_name = self.parse_programmer_name(programmer_name)

        # Check if programmer already exists (case-insensitive)
        existing = ProjectProgrammer.objects.filter(
            firstName__iexact=first_name,
            lastName__iexact=last_name
        ).first()

        if existing:
            return existing, f"Found existing programmer: {existing.firstName} {existing.lastName}"

        if dry_run:
            return None, f"Would create programmer: {first_name} {last_name}"

        # Use get_or_create with race condition protection
        try:
            with transaction.atomic():
                # Double-check within transaction to prevent race conditions
                existing = ProjectProgrammer.objects.filter(
                    firstName__iexact=first_name,
                    lastName__iexact=last_name
                ).first()

                if existing:
                    return existing, f"Found existing programmer (race condition): {existing.firstName} {existing.lastName}"

                # Try to find matching Person
                person = None
                if first_name and last_name:
                    # Try exact match first
                    person = Person.objects.filter(
                        firstName__iexact=first_name,
                        lastName__iexact=last_name
                    ).first()

                    # Try reverse order (lastName firstName)
                    if not person:
                        person = Person.objects.filter(
                            firstName__iexact=last_name,
                            lastName__iexact=first_name
                        ).first()

                # Create new ProjectProgrammer
                programmer = ProjectProgrammer.objects.create(
                    id=uuid.uuid4(),
                    firstName=first_name,
                    lastName=last_name,
                    person=person,
                    createdDate=timezone.now(),
                    updatedDate=timezone.now()
                )

                person_msg = f" (linked to Person: {person.firstName} {person.lastName})" if person else " (no Person link found)"
                return programmer, f"Created programmer: {programmer.firstName} {programmer.lastName}{person_msg}"

        except IntegrityError:
            # Race condition occurred, try to find the programmer that was created
            existing = ProjectProgrammer.objects.filter(
                firstName__iexact=first_name,
                lastName__iexact=last_name
            ).first()

            if existing:
                return existing, f"Found existing programmer (integrity error): {existing.firstName} {existing.lastName}"
            else:
                raise

    def _get_programmer_with_hierarchy(self, project_class):
        """
        Get programmer with hierarchical fallback logic.
        Uses shared utility with cycle detection for safety.
        """
        return get_programmer_from_hierarchy(project_class)

    def show_dry_run(self, specific_assignments, fallback_assignments, clear_existing, apply_to_projects):
        """Show what would be imported in dry-run mode"""
        self.stdout.write(self.style.SUCCESS("\n=== DRY RUN MODE ==="))

        if clear_existing:
            existing_count = ProjectClass.objects.exclude(defaultProgrammer=None).count()
            self.stdout.write(f"Would clear {existing_count} existing programmer assignments")

        if apply_to_projects:
            projects_to_update = Project.objects.filter(
                projectClass__defaultProgrammer__isnull=False,
                personProgramming__isnull=True
            ).count()
            self.stdout.write(f"Would apply default programmers to {projects_to_update} existing projects")

        # Show specific assignments
        if specific_assignments:
            self.stdout.write(f"\n--- SPECIFIC ASSIGNMENTS ({len(specific_assignments)}) ---")
            for i, assignment in enumerate(specific_assignments, 1):
                self.stdout.write(f"\n{i}. Class: {assignment['class_code']} - {assignment['class_description']}")
                self.stdout.write(f"   Programmer: {assignment['programmer_name']}")

                # Check if class exists
                project_class, class_msg = self.find_project_class(
                    assignment['class_code'],
                    assignment['class_description']
                )

                if project_class:
                    self.stdout.write(f"   Found class: {project_class.name[:60]}...")
                    if class_msg:
                        self.stdout.write(f"   WARNING: {class_msg}")
                else:
                    self.stdout.write(f"   ERROR: {class_msg}")

        # Show fallback assignments
        if fallback_assignments:
            self.stdout.write(f"\n--- FALLBACK ASSIGNMENTS ({len(fallback_assignments)}) ---")
            for i, assignment in enumerate(fallback_assignments, 1):
                self.stdout.write(f"\n{i}. District: {assignment['district_name']}")
                self.stdout.write(f"   Programmer: {assignment['programmer_name']}")

                # Find classes that would use this fallback
                classes = self.find_classes_for_district_fallback(assignment['district_name'])
                self.stdout.write(f"   Classes without specific programmer: {classes.count()}")

                if classes.count() > 0 and classes.count() <= 5:
                    for cls in classes:
                        self.stdout.write(f"     - {cls.name[:60]}...")

    @transaction.atomic
    def process_assignments(self, specific_assignments, fallback_assignments, clear_existing, apply_to_projects):
        """Process programmer assignments and assign them to project classes"""
        self.stdout.write(self.style.SUCCESS("\n=== IMPORTING PROGRAMMERS ==="))

        if clear_existing:
            cleared_count = ProjectClass.objects.exclude(defaultProgrammer=None).count()
            ProjectClass.objects.update(defaultProgrammer=None)
            self.stdout.write(f"Cleared {cleared_count} existing programmer assignments")

        created_programmers = 0
        found_programmers = 0
        assigned_classes = 0
        errors = []

        # Process specific assignments first
        self.stdout.write(f"\n--- PROCESSING SPECIFIC ASSIGNMENTS ({len(specific_assignments)}) ---")
        for i, assignment in enumerate(specific_assignments, 1):
            self.stdout.write(f"\n{i}/{len(specific_assignments)}: Processing {assignment['programmer_name']} -> {assignment['class_code']}")

            # Find project class
            project_class, class_msg = self.find_project_class(
                assignment['class_code'],
                assignment['class_description']
            )

            if not project_class:
                error_msg = f"Row {assignment['row']}: {class_msg}"
                errors.append(error_msg)
                self.stdout.write(f"   ERROR: {error_msg}")
                continue

            self.stdout.write(f"   Found class: {project_class.name[:60]}...")
            if class_msg:
                self.stdout.write(f"   WARNING: {class_msg}")

            # Find or create programmer
            try:
                programmer, prog_msg = self.find_or_create_programmer(assignment['programmer_name'])
                self.stdout.write(f"   {prog_msg}")

                if "Created" in prog_msg:
                    created_programmers += 1
                else:
                    found_programmers += 1

                # Assign programmer to class
                project_class.defaultProgrammer = programmer
                project_class.save()
                assigned_classes += 1
                self.stdout.write(f"   Assigned programmer to class")

            except Exception as e:
                error_msg = f"Row {assignment['row']}: Error processing programmer '{assignment['programmer_name']}': {e}"
                errors.append(error_msg)
                self.stdout.write(f"   ERROR: {error_msg}")

        # Process fallback assignments
        self.stdout.write(f"\n--- PROCESSING FALLBACK ASSIGNMENTS ({len(fallback_assignments)}) ---")
        for i, assignment in enumerate(fallback_assignments, 1):
            self.stdout.write(f"\n{i}/{len(fallback_assignments)}: Processing fallback {assignment['programmer_name']} -> {assignment['district_name']}")

            # Find classes that need this fallback
            classes = self.find_classes_for_district_fallback(assignment['district_name'])
            self.stdout.write(f"   Found {classes.count()} classes without specific programmer in {assignment['district_name']}")

            if classes.count() == 0:
                self.stdout.write(f"   WARNING: No classes found or all already have programmers assigned")
                continue

            # Find or create programmer
            try:
                programmer, prog_msg = self.find_or_create_programmer(assignment['programmer_name'])
                self.stdout.write(f"   {prog_msg}")

                if "Created" in prog_msg:
                    created_programmers += 1
                else:
                    found_programmers += 1

                # Assign programmer to all classes in this district that don't have one
                # Use bulk update for better performance
                fallback_assigned = classes.update(defaultProgrammer=programmer)
                assigned_classes += fallback_assigned

                self.stdout.write(f"   Assigned fallback programmer to {fallback_assigned} classes")

            except Exception as e:
                error_msg = f"Row {assignment['row']}: Error processing fallback programmer '{assignment['programmer_name']}': {e}"
                errors.append(error_msg)
                self.stdout.write(f"   ERROR: {error_msg}")

        # Apply default programmers to existing projects if requested
        projects_updated = 0
        if apply_to_projects:
            self.stdout.write(f"\n--- APPLYING DEFAULT PROGRAMMERS TO EXISTING PROJECTS ---")

            # Find all projects that:
            # 1. Have a projectClass
            # 2. Don't already have a personProgramming assigned
            # Note: We use hierarchical fallback in Python to find programmers
            projects_to_update = Project.objects.filter(
                personProgramming__isnull=True,
                projectClass__isnull=False
            ).select_related('projectClass')

            projects_count = projects_to_update.count()
            self.stdout.write(f"Found {projects_count} projects without programmers")

            # Use hierarchical fallback to find programmers
            for project in projects_to_update:
                programmer = self._get_programmer_with_hierarchy(project.projectClass)
                if programmer:
                    project.personProgramming = programmer
                    project.save()
                    projects_updated += 1

            self.stdout.write(f"Updated {projects_updated} projects with default programmers (including hierarchical fallback)")

        # Summary
        self.stdout.write(self.style.SUCCESS(f"\n=== IMPORT SUMMARY ==="))
        self.stdout.write(f"Created programmers: {created_programmers}")
        self.stdout.write(f"Found existing programmers: {found_programmers}")
        self.stdout.write(f"Total classes assigned: {assigned_classes}")
        if apply_to_projects:
            self.stdout.write(f"Projects updated with programmers: {projects_updated}")

        if errors:
            self.stdout.write(f"Errors: {len(errors)}")
            for error in errors:
                self.stdout.write(f"   {error}")
        else:
            self.stdout.write("No errors!")

        self.stdout.write(self.style.SUCCESS("\nProgrammer import completed successfully!"))
